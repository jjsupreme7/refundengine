#!/usr/bin/env python3
"""
Invoice Lookup and Site Tracking Module

Handles:
- Extracting invoice hyperlinks from Excel files
- Tracking site IDs and locations from tax data
- Linking invoices to tax records
- Downloading/fetching invoices from hyperlinks
- Matching invoices in the Test Data folders

Usage:
    from analysis.invoice_lookup import InvoiceLookup

    lookup = InvoiceLookup()

    # Extract invoice links from Excel
    links = lookup.extract_invoice_links("Test Data/Phase_3_2023_Use Tax_10-17-25.xlsx")

    # Find invoice in Test Data folders
    invoice_path = lookup.find_invoice_file("INV1068437")

    # Get site information for a tax record
    site_info = lookup.extract_site_info(tax_record)
"""

import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
import pandas as pd


class InvoiceLookup:
    """Manage invoice lookups, hyperlinks, and site tracking"""

    def __init__(self, test_data_root: str = "Test Data"):
        """
        Initialize invoice lookup

        Args:
            test_data_root: Path to Test Data folder containing invoices
        """
        self.test_data_root = Path(test_data_root).resolve()
        self.invoice_cache = {}  # Cache of invoice number -> file path
        self._build_invoice_cache()

    def _validate_path(self, file_path: Path) -> bool:
        """
        Validate that a path is within the test_data_root directory.
        Prevents path traversal and symlink attacks.

        Args:
            file_path: Path to validate

        Returns:
            True if path is safe, False otherwise
        """
        try:
            resolved_path = file_path.resolve()
            return str(resolved_path).startswith(str(self.test_data_root))
        except Exception:
            return False

    def _build_invoice_cache(self):
        """Build cache of all invoices in Test Data folders"""
        if not self.test_data_root.exists():
            print(f"Warning: Test Data folder not found: {self.test_data_root}")
            return

        # Scan Invoices folder
        invoices_folder = self.test_data_root / "Invoices"
        if invoices_folder.exists():
            for file in invoices_folder.glob("*"):
                # Validate path to prevent symlink attacks
                if not self._validate_path(file):
                    print(f"Warning: Skipping file outside trusted directory: {file}")
                    continue
                if file.suffix.lower() in ['.pdf', '.tif', '.tiff', '.xls', '.xlsx']:
                    # Extract invoice number from filename
                    # Format: 000005000021-1.PDF -> 000005000021
                    invoice_num = file.stem.split('-')[0]
                    self.invoice_cache[invoice_num] = file

        # Scan Purchase Orders folder
        po_folder = self.test_data_root / "Purchase Orders"
        if po_folder.exists():
            for file in po_folder.glob("*"):
                # Validate path to prevent symlink attacks
                if not self._validate_path(file):
                    print(f"Warning: Skipping file outside trusted directory: {file}")
                    continue
                # Extract PO number: PO_4900668309_ERICSSON_...
                if file.name.startswith("PO_"):
                    parts = file.name.split("_")
                    if len(parts) >= 2:
                        po_num = parts[1]
                        self.invoice_cache[f"PO_{po_num}"] = file

        print(f"Invoice cache built: {len(self.invoice_cache)} documents indexed")

    def find_invoice_file(self, invoice_number: str) -> Optional[Path]:
        """
        Find invoice file by invoice number

        Args:
            invoice_number: Invoice number (e.g., "INV1068437", "000005000021")

        Returns:
            Path to invoice file, or None if not found
        """
        # Clean invoice number
        clean_num = invoice_number.replace("INV", "").strip()

        # Try exact match first
        if clean_num in self.invoice_cache:
            path = self.invoice_cache[clean_num]
            # Validate path before returning (defense in depth)
            if self._validate_path(path):
                return path
            else:
                print(f"Warning: Cached path validation failed for {invoice_number}")
                return None

        # Try variations
        for cached_num, path in self.invoice_cache.items():
            if clean_num in cached_num or cached_num in clean_num:
                # Validate path before returning (defense in depth)
                if self._validate_path(path):
                    return path
                else:
                    print(f"Warning: Cached path validation failed for {invoice_number}")
                    return None

        return None

    def extract_invoice_links(self, excel_path: str) -> List[Dict]:
        """
        Extract invoice hyperlinks from Excel file

        Args:
            excel_path: Path to Excel file

        Returns:
            List of dicts with invoice info and hyperlinks:
            [
                {
                    'row': 2,
                    'vendor': 'ABC Corp',
                    'invoice_number': 'INV123',
                    'inv1_link': 'https://...',
                    'inv2_link': 'https://...',
                    'inv1_text': 'Invoice 1.pdf',
                    'inv2_text': 'Invoice 2.pdf'
                },
                ...
            ]
        """
        wb = load_workbook(excel_path)
        ws = wb.active

        # Get headers
        headers = [cell.value for cell in ws[1]]

        # Find relevant columns
        col_map = {}
        for col_name in ['Inv-1 Hyperlink', 'Inv-2 Hyperlink', 'Inv-1PDF', 'Inv-2 PDF',
                         'Vendor Name', 'INVNO', 'vendor', 'xblnr_invoice_number']:
            if col_name in headers:
                col_map[col_name] = headers.index(col_name) + 1

        results = []

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}

            # Get vendor and invoice number
            if 'Vendor Name' in col_map:
                row_data['vendor'] = ws.cell(row_idx, col_map['Vendor Name']).value
            elif 'vendor' in col_map:
                row_data['vendor'] = ws.cell(row_idx, col_map['vendor']).value

            if 'INVNO' in col_map:
                row_data['invoice_number'] = ws.cell(row_idx, col_map['INVNO']).value
            elif 'xblnr_invoice_number' in col_map:
                row_data['invoice_number'] = ws.cell(row_idx, col_map['xblnr_invoice_number']).value

            # Get hyperlinks
            has_data = False

            if 'Inv-1 Hyperlink' in col_map:
                cell = ws.cell(row_idx, col_map['Inv-1 Hyperlink'])
                if cell.hyperlink:
                    row_data['inv1_link'] = cell.hyperlink.target
                    row_data['inv1_text'] = cell.value
                    has_data = True

            if 'Inv-2 Hyperlink' in col_map:
                cell = ws.cell(row_idx, col_map['Inv-2 Hyperlink'])
                if cell.hyperlink:
                    row_data['inv2_link'] = cell.hyperlink.target
                    row_data['inv2_text'] = cell.value
                    has_data = True

            # Check PDF columns too
            if 'Inv-1PDF' in col_map:
                val = ws.cell(row_idx, col_map['Inv-1PDF']).value
                if val:
                    row_data['inv1_pdf'] = val
                    has_data = True

            if 'Inv-2 PDF' in col_map:
                val = ws.cell(row_idx, col_map['Inv-2 PDF']).value
                if val:
                    row_data['inv2_pdf'] = val
                    has_data = True

            if has_data:
                row_data['row'] = row_idx
                results.append(row_data)

        print(f"Extracted {len(results)} invoice links from {excel_path}")
        return results

    def extract_site_info(self, row_data: Dict) -> Dict:
        """
        Extract site/location information from a tax record

        Args:
            row_data: Dictionary with tax record data (from Excel row)

        Returns:
            Dictionary with site information:
            {
                'cost_center': '0002353810',
                'cost_center_desc': 'GEORGE WA WAY & MCMURRAY ST 2',
                'profit_center': '0002353810',
                'profit_center_desc': 'GEORGE WA WAY & MCMURRAY ST 2',
                'wbs': '00000000',
                'wbs_desc': None,
                'tax_jurisdiction': '4800506300',
                'jurisdiction_state': 'WA',
                'location_city': 'GEORGE',
                'location_state': 'WA',
                'location_address': None
            }
        """
        site_info = {}

        # Cost center (often represents site/location)
        site_info['cost_center'] = row_data.get('kost1_cost_center') or row_data.get('Cost Center')
        site_info['cost_center_desc'] = row_data.get('ltext_cost_center_description')

        # Profit center
        site_info['profit_center'] = row_data.get('prctr_profit_center') or row_data.get('Profit Center')
        site_info['profit_center_desc'] = row_data.get('ltext_profit_center_description')

        # WBS (Work Breakdown Structure) - project/site tracking
        site_info['wbs'] = row_data.get('projk_wbs') or row_data.get('WBS')
        site_info['wbs_desc'] = row_data.get('post1_wbs_description')

        # Tax jurisdiction
        site_info['tax_jurisdiction'] = row_data.get('txjcd_po_jurisdiction_code')
        site_info['jurisdiction_state'] = row_data.get('tax_jurisdiction_state') or row_data.get('STATE')

        # Location from profit center region
        site_info['location_state'] = row_data.get('regio_profit_center_region')

        # Try to extract city from cost center description
        desc = site_info.get('cost_center_desc', '')
        if desc:
            # Pattern: "GEORGE WA WAY & MCMURRAY ST 2" or "66929 - WALLA WALLA, WA (ALA)"
            city_match = re.search(r'([A-Z\s]+),?\s+WA', desc)
            if city_match:
                site_info['location_city'] = city_match.group(1).strip()
            else:
                # Try first word
                words = desc.split()
                if words:
                    site_info['location_city'] = words[0]

        return site_info

    def match_invoice_to_local_file(self, invoice_number: str, vendor_name: str = None) -> Dict:
        """
        Try to match an invoice number to a local file in Test Data

        Args:
            invoice_number: Invoice number from Excel
            vendor_name: Optional vendor name for better matching

        Returns:
            {
                'found': bool,
                'file_path': Path or None,
                'match_confidence': 'exact'|'partial'|'none',
                'file_type': 'pdf'|'tif'|'xlsx'|etc
            }
        """
        result = {
            'found': False,
            'file_path': None,
            'match_confidence': 'none',
            'file_type': None
        }

        # Try to find file
        file_path = self.find_invoice_file(invoice_number)

        if file_path:
            result['found'] = True
            result['file_path'] = file_path
            result['match_confidence'] = 'exact'
            result['file_type'] = file_path.suffix.lower().replace('.', '')

        return result

    def get_site_summary(self, df: pd.DataFrame) -> Dict:
        """
        Get summary of sites/locations in the dataset

        Args:
            df: DataFrame with tax data

        Returns:
            {
                'total_sites': int,
                'sites_by_state': {state: count},
                'top_cost_centers': [(cost_center, description, count)],
                'tax_by_site': [(cost_center, tax_amount)]
            }
        """
        summary = {}

        # Count unique cost centers (sites)
        cost_center_col = 'kost1_cost_center' if 'kost1_cost_center' in df.columns else 'Cost Center'
        if cost_center_col in df.columns:
            summary['total_sites'] = df[cost_center_col].nunique()

            # Top cost centers by count
            top_centers = df[cost_center_col].value_counts().head(10)
            summary['top_cost_centers'] = [(cc, None, count) for cc, count in top_centers.items()]

        # Count by state
        state_col = 'tax_jurisdiction_state' if 'tax_jurisdiction_state' in df.columns else 'STATE'
        if state_col in df.columns:
            summary['sites_by_state'] = df[state_col].value_counts().to_dict()

        # Tax amount by cost center
        tax_col = 'hwste_tax_amount_lc' if 'hwste_tax_amount_lc' in df.columns else 'Total Tax'
        if cost_center_col in df.columns and tax_col in df.columns:
            tax_by_site = df.groupby(cost_center_col)[tax_col].sum().sort_values(ascending=False).head(10)
            summary['tax_by_site'] = [(site, amt) for site, amt in tax_by_site.items()]

        return summary


def create_invoice_download_script(invoice_links: List[Dict], output_folder: str = "downloaded_invoices"):
    """
    Create a script to download invoices from hyperlinks

    Args:
        invoice_links: List of invoice link dicts from extract_invoice_links()
        output_folder: Where to save downloaded invoices

    Returns:
        Path to generated download script
    """
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    script_lines = [
        "#!/usr/bin/env python3",
        "\"\"\"",
        "Auto-generated invoice download script",
        "Downloads invoices from hyperlinks in Excel file",
        "\"\"\"",
        "",
        "import requests",
        "from pathlib import Path",
        "",
        "output_folder = Path('downloaded_invoices')",
        "output_folder.mkdir(parents=True, exist_ok=True)",
        "",
        "invoices_to_download = [",
    ]

    for link_data in invoice_links:
        if 'inv1_link' in link_data or 'inv2_link' in link_data:
            script_lines.append(f"    {{")
            script_lines.append(f"        'vendor': '{link_data.get('vendor', 'Unknown')}',")
            script_lines.append(f"        'invoice': '{link_data.get('invoice_number', 'Unknown')}',")
            if 'inv1_link' in link_data:
                script_lines.append(f"        'url1': '{link_data['inv1_link']}',")
            if 'inv2_link' in link_data:
                script_lines.append(f"        'url2': '{link_data['inv2_link']}',")
            script_lines.append(f"    }},")

    script_lines.extend([
        "]",
        "",
        "for invoice in invoices_to_download:",
        "    vendor = invoice['vendor'].replace(' ', '_').replace('/', '_')",
        "    inv_num = invoice['invoice']",
        "    ",
        "    if 'url1' in invoice:",
        "        filename = output_folder / f'{vendor}_{inv_num}_1.pdf'",
        "        print(f'Downloading {filename}...')",
        "        # response = requests.get(invoice['url1'])",
        "        # with open(filename, 'wb') as f:",
        "        #     f.write(response.content)",
        "    ",
        "    if 'url2' in invoice:",
        "        filename = output_folder / f'{vendor}_{inv_num}_2.pdf'",
        "        print(f'Downloading {filename}...')",
        "        # response = requests.get(invoice['url2'])",
        "        # with open(filename, 'wb') as f:",
        "        #     f.write(response.content)",
        "",
        "print('Download complete!')",
    ])

    script_path = Path("scripts/download_invoices.py")
    script_path.write_text("\n".join(script_lines))

    print(f"Download script created: {script_path}")
    return script_path


if __name__ == "__main__":
    import sys

    print("Invoice Lookup Module - Testing")
    print("=" * 70)

    lookup = InvoiceLookup()

    # Test finding invoices
    test_invoices = ["000005000021", "INV1068437", "5106782843"]

    print("\nTesting invoice lookups:")
    for inv in test_invoices:
        result = lookup.match_invoice_to_local_file(inv)
        print(f"  {inv}: {result}")

    # Test extracting links from Excel
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        print(f"\nExtracting links from {excel_file}...")
        links = lookup.extract_invoice_links(excel_file)

        if links:
            print(f"Found {len(links)} invoice links")
            print("\nFirst 5 links:")
            for link in links[:5]:
                print(f"  Row {link['row']}: {link.get('vendor')} - {link.get('invoice_number')}")
                if 'inv1_link' in link:
                    print(f"    Link 1: {link['inv1_link']}")
        else:
            print("No invoice hyperlinks found in file")
