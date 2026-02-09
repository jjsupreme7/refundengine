#!/usr/bin/env python3
"""Extract vendor profiles from the Phase 2 Master Refunds file.

Usage:
    python scripts/extract_vendor_profiles.py [path_to_master_file]

Defaults to ~/Downloads/Phase 2 Master Refunds Jun 15 2025.xlsx
Writes to config/vendor_profiles.json
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


def _select_examples(rows: list[dict], max_examples: int = 3) -> list[dict]:
    """Select diverse few-shot examples, preferring rows with analyst notes."""
    if not rows:
        return []
    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        key = (row["tax_category"], row["refund_basis"])
        groups[key].append(row)
    sorted_groups = sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)
    examples = []
    for _key, group_rows in sorted_groups:
        if len(examples) >= max_examples:
            break
        best = max(group_rows, key=lambda r: len(r.get("notes", "")))
        examples.append(best)
    return examples


def _dominant(counter: Counter) -> dict:
    if not counter:
        return {"value": "Unknown", "count": 0}
    value, count = counter.most_common(1)[0]
    return {"value": str(value), "count": count}


def extract_profiles(master_path: str, output_path: str, min_rows: int = 3) -> None:
    print(f"Reading {master_path}...")
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb["Refund Summary"]

    # Column indices (1-based): A=1 Vendor, C=3 Claim, F=6 Notes,
    # T=20 Sales/Use, X=24 Non-Taxable%, Y=25 Methodology,
    # AE=31 Type, AG=33 Description, AI=35 Tax Category, AK=37 Refund Basis
    vendors: dict[str, dict] = {}

    for r in range(2, ws.max_row + 1):
        vendor = ws.cell(r, 1).value
        if not vendor:
            continue
        vendor = str(vendor).strip().upper()

        if vendor not in vendors:
            vendors[vendor] = {
                "total_rows": 0,
                "claimed_count": 0,
                "pass_count": 0,
                "tax_categories": Counter(),
                "product_types": Counter(),
                "refund_bases": Counter(),
                "methodologies": Counter(),
                "methodology_pcts": defaultdict(list),
                "descriptions": Counter(),
                "example_rows": [],
            }

        v = vendors[vendor]
        v["total_rows"] += 1

        claim = ws.cell(r, 3).value
        if claim and "Y" in str(claim):
            v["claimed_count"] += 1

        notes = ws.cell(r, 6).value
        if notes and "Pass" in str(notes):
            v["pass_count"] += 1

        tax_cat = ws.cell(r, 35).value
        if tax_cat:
            v["tax_categories"][str(tax_cat)] += 1

        ptype = ws.cell(r, 31).value
        if ptype:
            v["product_types"][str(ptype)] += 1

        rbasis = ws.cell(r, 37).value
        if rbasis:
            v["refund_bases"][str(rbasis)] += 1

        meth = ws.cell(r, 25).value
        if meth:
            meth_str = str(meth)
            v["methodologies"][meth_str] += 1
            # Capture allocation percentage per methodology
            pct = ws.cell(r, 24).value
            if pct is not None:
                try:
                    pct_val = float(pct)
                    if 0.0 <= pct_val <= 1.0:
                        v["methodology_pcts"][meth_str].append(pct_val)
                except (TypeError, ValueError):
                    pass

        desc = ws.cell(r, 33).value
        if desc and len(str(desc).strip()) > 5:
            v["descriptions"][str(desc).strip()[:80]] += 1

        # Collect row data for few-shot example selection
        if tax_cat:
            notes_text = str(notes).strip()[:120] if notes else ""
            v["example_rows"].append({
                "description": str(desc).strip()[:80] if desc else "",
                "tax_category": str(tax_cat),
                "refund_basis": str(rbasis) if rbasis else "",
                "methodology": str(meth) if meth else "",
                "product_type": str(ptype) if ptype else "",
                "notes": notes_text,
            })

    # Filter to vendors with min_rows classified rows
    output_vendors = {}
    for name, v in sorted(vendors.items()):
        classified = sum(v["tax_categories"].values())
        if classified < min_rows:
            continue

        # Build methodology_mix with allocation percentages
        methodology_mix = {}
        for meth_name, count in v["methodologies"].most_common():
            pcts = v["methodology_pcts"].get(meth_name, [])
            avg_pct = round(sum(pcts) / len(pcts), 4) if pcts else None
            methodology_mix[meth_name] = {"count": count, "avg_pct": avg_pct}

        output_vendors[name] = {
            "total_rows": v["total_rows"],
            "claimed_count": v["claimed_count"],
            "pass_count": v["pass_count"],
            "dominant_tax_category": _dominant(v["tax_categories"]),
            "dominant_product_type": _dominant(v["product_types"]),
            "dominant_refund_basis": _dominant(v["refund_bases"]),
            "dominant_methodology": _dominant(v["methodologies"]),
            "all_tax_categories": dict(v["tax_categories"].most_common()),
            "all_refund_bases": dict(v["refund_bases"].most_common()),
            "all_methodologies": dict(v["methodologies"].most_common()),
            "methodology_mix": methodology_mix,
            "sample_descriptions": [
                d for d, _ in v["descriptions"].most_common(3)
            ],
            "few_shot_examples": _select_examples(v["example_rows"]),
        }

    result = {
        "_metadata": {
            "source": str(master_path),
            "extracted_at": datetime.now().isoformat(),
            "total_vendors": len(output_vendors),
            "min_classified_rows": min_rows,
        },
        "vendors": output_vendors,
    }

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(result, f, indent=2)

    print(f"Wrote {len(output_vendors)} vendor profiles to {output_path}")


if __name__ == "__main__":
    default_master = str(
        Path.home() / "Downloads" / "Phase 2 Master Refunds Jun 15 2025.xlsx"
    )
    master = sys.argv[1] if len(sys.argv) > 1 else default_master
    output = str(Path(__file__).resolve().parent.parent / "config" / "vendor_profiles.json")
    extract_profiles(master, output)
