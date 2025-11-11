#!/usr/bin/env python3
"""
Import Knowledge Base Metadata from Excel

Detects changes between Excel file and current database, shows diff, and updates Supabase.
Only updates changed fields, preserves embeddings and chunk text.

Usage:
    python scripts/import_metadata_excel.py --file metadata.xlsx
    python scripts/import_metadata_excel.py --file metadata.xlsx --auto-confirm
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any
from supabase import create_client
from dotenv import load_dotenv
import argparse
from openpyxl import load_workbook

# Load environment
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


class ExcelMetadataImporter:
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        self.changes = []

    def load_sheet_as_dicts(self, sheet_name: str) -> List[Dict]:
        """Load Excel sheet and convert to list of dicts"""
        print(f"📖 Reading sheet: {sheet_name}")

        wb = load_workbook(self.excel_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Sheet '{sheet_name}' not found in Excel file")
            return []

        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue

            row_dict = {}
            for header, value in zip(headers, row):
                # Convert empty cells to None
                if value == "" or value == "None":
                    row_dict[header] = None
                else:
                    row_dict[header] = value

            rows.append(row_dict)

        print(f"✅ Loaded {len(rows)} rows from '{sheet_name}'")
        return rows

    def fetch_current_data(self, table_name: str, ids: List[str]) -> Dict[str, Dict]:
        """Fetch current data from Supabase for given IDs"""
        print(f"📥 Fetching current data from {table_name}...")

        # Fetch in batches to avoid URL length limits
        batch_size = 50
        all_data = []

        for i in range(0, len(ids), batch_size):
            batch_ids = ids[i : i + batch_size]
            result = (
                self.supabase.table(table_name)
                .select("*")
                .in_("id", batch_ids)
                .execute()
            )
            all_data.extend(result.data)

        # Convert to dict keyed by ID
        current_data = {row["id"]: row for row in all_data}
        print(f"✅ Fetched {len(current_data)} records from database")

        return current_data

    def detect_changes(
        self, excel_rows: List[Dict], current_data: Dict[str, Dict], table_name: str
    ) -> List[Dict]:
        """Compare Excel rows with current database and detect changes"""
        print(f"\n🔍 Detecting changes in {table_name}...")

        changes = []

        for excel_row in excel_rows:
            row_id = excel_row.get("id")
            if not row_id:
                print(f"⚠️  Skipping row without ID: {excel_row}")
                continue

            current_row = current_data.get(row_id)
            if not current_row:
                print(f"⚠️  Warning: ID {row_id} not found in database (skipping)")
                continue

            # Compare each field
            changed_fields = {}
            for field, excel_value in excel_row.items():
                # Skip ID fields and non-editable fields
                if field in [
                    "id",
                    "document_id",
                    "created_at",
                    "updated_at",
                    "embedding",
                    "chunk_text",
                ]:
                    continue

                current_value = current_row.get(field)

                # Normalize for comparison
                excel_val_normalized = self._normalize_value(excel_value)
                current_val_normalized = self._normalize_value(current_value)

                if excel_val_normalized != current_val_normalized:
                    changed_fields[field] = {"old": current_value, "new": excel_value}

            if changed_fields:
                changes.append(
                    {
                        "id": row_id,
                        "table": table_name,
                        "changed_fields": changed_fields,
                        "excel_row": excel_row,
                        "current_row": current_row,
                    }
                )

        print(f"✅ Found {len(changes)} records with changes")
        return changes

    def _normalize_value(self, value: Any) -> Any:
        """Normalize values for comparison"""
        if value is None or value == "" or value == "None":
            return None

        # Handle dates
        if hasattr(value, "date"):
            return str(value.date())

        # Handle arrays - convert both database arrays and Excel strings to sorted comma-separated strings
        if isinstance(value, list):
            # Database array: ['sales tax', 'use tax'] → 'sales tax, use tax'
            return ", ".join(sorted(value)) if value else None

        # Excel string might already be comma-separated
        if isinstance(value, str) and "," in value:
            # Normalize: split, strip, sort, rejoin
            items = [item.strip() for item in value.split(",") if item.strip()]
            return ", ".join(sorted(items)) if items else None

        return str(value).strip()

    def show_diff(self, changes: List[Dict]):
        """Display a human-readable diff of changes"""
        print("\n" + "=" * 80)
        print("📊 CHANGES DETECTED")
        print("=" * 80)

        for i, change in enumerate(changes, 1):
            record_id = change["id"]
            table = change["table"]
            changed_fields = change["changed_fields"]

            # Get a display name for the record
            excel_row = change["excel_row"]
            display_name = (
                excel_row.get("title")
                or excel_row.get("citation")
                or excel_row.get("vendor_name")
                or record_id[:8]
            )

            print(f"\n[{i}/{len(changes)}] {table}: {display_name}")
            print(f"    ID: {record_id}")

            for field, values in changed_fields.items():
                old_val = values["old"]
                new_val = values["new"]

                print(f"\n    Field: {field}")
                print(f"      - Old: {old_val}")
                print(f"      + New: {new_val}")

        print("\n" + "=" * 80)
        print(f"Total changes: {len(changes)} records")
        print("=" * 80)

    def confirm_changes(self) -> bool:
        """Ask user to confirm changes"""
        response = (
            input("\n❓ Apply these changes to Supabase? [y/N]: ").strip().lower()
        )
        return response in ["y", "yes"]

    def _prepare_value_for_db(self, field_name: str, value: Any) -> Any:
        """Convert Excel values to database format"""
        # Array fields that need to be converted from strings to arrays
        ARRAY_FIELDS = ["topic_tags", "tax_types", "industries", "referenced_statutes"]

        if field_name in ARRAY_FIELDS:
            if isinstance(value, str):
                # Convert comma-separated string to array
                items = [item.strip() for item in value.split(",") if item.strip()]
                return items if items else []
            elif isinstance(value, list):
                return value
            else:
                return []

        return value

    def apply_changes(self, changes: List[Dict]) -> Tuple[int, int]:
        """Apply changes to Supabase with cascading updates to chunks"""
        print("\n🔄 Applying changes to Supabase...")

        success_count = 0
        error_count = 0
        cascade_count = 0

        # Fields that should cascade from documents to chunks
        CASCADE_FIELDS = [
            "citation",
            "law_category",
            "topic_tags",
            "tax_types",
            "industries",
            "referenced_statutes",
            "vendor_name",
            "vendor_category",
        ]

        for change in changes:
            record_id = change["id"]
            table = change["table"]
            changed_fields = change["changed_fields"]

            # Build update dict with only changed fields - convert to DB format
            update_data = {}
            for field, values in changed_fields.items():
                update_data[field] = self._prepare_value_for_db(field, values["new"])

            # Add updated_at timestamp
            update_data["updated_at"] = datetime.now().isoformat()

            try:
                result = (
                    self.supabase.table(table)
                    .update(update_data)
                    .eq("id", record_id)
                    .execute()
                )

                if result.data:
                    success_count += 1
                    display_name = (
                        change["excel_row"].get("title")
                        or change["excel_row"].get("citation")
                        or record_id[:8]
                    )
                    print(f"  ✅ Updated: {display_name}")

                    # CASCADE: If this is a document change, update its chunks too
                    if table == "knowledge_documents":
                        cascade_fields = {
                            k: v for k, v in update_data.items() if k in CASCADE_FIELDS
                        }

                        if cascade_fields:
                            # Determine chunk table based on document type
                            doc_type = change["excel_row"].get("document_type")
                            chunk_table = (
                                "tax_law_chunks"
                                if doc_type == "tax_law"
                                else "vendor_background_chunks"
                            )

                            try:
                                # Update all chunks for this document
                                chunk_result = (
                                    self.supabase.table(chunk_table)
                                    .update(cascade_fields)
                                    .eq("document_id", record_id)
                                    .execute()
                                )

                                if chunk_result.data:
                                    num_chunks = len(chunk_result.data)
                                    cascade_count += num_chunks
                                    print(f"    ↳ Cascaded to {num_chunks} chunks")

                            except Exception as e:
                                print(f"    ⚠️  Could not cascade to chunks: {e}")

                else:
                    error_count += 1
                    print(f"  ❌ Failed to update: {record_id}")

            except Exception as e:
                error_count += 1
                print(f"  ❌ Error updating {record_id}: {e}")

        if cascade_count > 0:
            print(f"\n  🔄 Cascaded changes to {cascade_count} chunks")

        return success_count, error_count

    def run(self, auto_confirm: bool = False):
        """Main import workflow"""
        print("=" * 80)
        print("📥 METADATA IMPORT FROM EXCEL")
        print("=" * 80)
        print(f"Excel File: {self.excel_path}")

        all_changes = []

        # ========================================================================
        # Process Documents sheet
        # ========================================================================
        try:
            docs_rows = self.load_sheet_as_dicts("Documents")
            if docs_rows:
                ids = [row["id"] for row in docs_rows if row.get("id")]
                if ids:
                    current_data = self.fetch_current_data("knowledge_documents", ids)
                    changes = self.detect_changes(
                        docs_rows, current_data, "knowledge_documents"
                    )
                    all_changes.extend(changes)
        except Exception as e:
            print(f"⚠️  Could not process Documents sheet: {e}")

        # ========================================================================
        # Process Tax Law Chunks sheet
        # ========================================================================
        try:
            chunks_rows = self.load_sheet_as_dicts("Tax Law Chunks")
            if chunks_rows:
                ids = [row["id"] for row in chunks_rows if row.get("id")]
                if ids:
                    current_data = self.fetch_current_data("tax_law_chunks", ids)
                    changes = self.detect_changes(
                        chunks_rows, current_data, "tax_law_chunks"
                    )
                    all_changes.extend(changes)
        except Exception as e:
            print(f"⚠️  Could not process Tax Law Chunks sheet: {e}")

        # ========================================================================
        # Process Vendor Chunks sheet
        # ========================================================================
        try:
            vendor_rows = self.load_sheet_as_dicts("Vendor Chunks")
            if vendor_rows:
                ids = [row["id"] for row in vendor_rows if row.get("id")]
                if ids:
                    current_data = self.fetch_current_data(
                        "vendor_background_chunks", ids
                    )
                    changes = self.detect_changes(
                        vendor_rows, current_data, "vendor_background_chunks"
                    )
                    all_changes.extend(changes)
        except Exception as e:
            print(f"⚠️  Could not process Vendor Chunks sheet: {e}")

        # ========================================================================
        # Show changes and apply
        # ========================================================================
        if not all_changes:
            print("\n✅ No changes detected. Database is already up to date!")
            return 0

        # Show diff
        self.show_diff(all_changes)

        # Confirm changes
        if not auto_confirm:
            if not self.confirm_changes():
                print("\n❌ Import cancelled by user")
                return 1

        # Apply changes
        success_count, error_count = self.apply_changes(all_changes)

        # Summary
        print("\n" + "=" * 80)
        print("✅ IMPORT COMPLETE")
        print("=" * 80)
        print(f"Successfully updated: {success_count} records")
        if error_count > 0:
            print(f"Errors: {error_count} records")
        print("=" * 80)

        return 0 if error_count == 0 else 1


def main():
    parser = argparse.ArgumentParser(
        description="Import metadata changes from Excel to Supabase"
    )
    parser.add_argument(
        "--file", type=str, required=True, help="Path to edited Excel file"
    )
    parser.add_argument(
        "--auto-confirm",
        action="store_true",
        help="Skip confirmation prompt and apply changes automatically",
    )

    args = parser.parse_args()

    # Validate file exists
    excel_path = Path(args.file)
    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        return 1

    if not excel_path.suffix in [".xlsx", ".xlsm"]:
        print(f"❌ File must be Excel format (.xlsx or .xlsm): {excel_path}")
        return 1

    try:
        importer = ExcelMetadataImporter(str(excel_path))
        return importer.run(auto_confirm=args.auto_confirm)

    except Exception as e:
        print(f"\n❌ Error during import: {e}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
