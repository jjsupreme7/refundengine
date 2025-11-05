#!/usr/bin/env python3
"""
Internal Analysis Workbook Generator

Creates comprehensive internal analysis workbook with 6 sheets.

Usage:
    python scripts/generate_internal_workbook.py --client_id 1
"""

import argparse
import sys
import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
import excel_utils as xl

def get_database_connection():
    project_root = Path(__file__).parent.parent
    db_path = project_root / "database" / "refund_engine.db"
    return sqlite3.connect(str(db_path))

def generate_internal_analysis_workbook(client_id):
    """Generate comprehensive internal analysis workbook."""
    conn = get_database_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT client_name FROM clients WHERE id = ?", (client_id,))
    client = cursor.fetchone()
    if not client:
        print(f"❌ Client {client_id} not found")
        return None

    client_name = client[0]
    print(f"\n📊 Generating internal analysis workbook for: {client_name}")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Dashboard
    ws1 = wb.create_sheet("Dashboard")
    xl.add_title(ws1, f"{client_name} - Internal Analysis Dashboard", row=1, end_col=6)

    cursor.execute("""
        SELECT COUNT(*), SUM(estimated_refund_amount), AVG(confidence_score)
        FROM refund_analysis ra
        JOIN invoices i ON ra.invoice_id = i.id
        WHERE i.client_id = ? AND ra.potentially_eligible = 1
    """, (client_id,))

    count, total, avg = cursor.fetchone()
    ws1.cell(3, 1, "Total Refund Opportunities:").font = xl.Font(bold=True)
    ws1.cell(3, 2, count or 0)
    ws1.cell(4, 1, "Total Refund Amount:").font = xl.Font(bold=True)
    ws1.cell(4, 2, f"${total or 0:.2f}")
    ws1.cell(5, 1, "Average Confidence:").font = xl.Font(bold=True)
    ws1.cell(5, 2, f"{avg or 0:.1f}%")

    xl.auto_adjust_columns(ws1)

    # Sheet 2: All Invoices
    ws2 = wb.create_sheet("All Invoices")
    headers = ["Invoice #", "Date", "Vendor", "Total", "Tax", "Status"]
    xl.add_header_row(ws2, headers)

    cursor.execute("""
        SELECT invoice_number, invoice_date, vendor_name, total_amount, sales_tax_charged, 'Analyzed'
        FROM invoices WHERE client_id = ?
    """, (client_id,))

    for i, row in enumerate(cursor.fetchall(), 2):
        for j, val in enumerate(row, 1):
            ws2.cell(i, j, val)

    xl.format_as_table(ws2)

    # Sheet 3: High Confidence (>80%)
    ws3 = wb.create_sheet("High Confidence (>80%)")
    xl.add_header_row(ws3, ["Description", "Refund", "Confidence", "Exemption"])

    cursor.execute("""
        SELECT li.item_description, ra.estimated_refund_amount, ra.confidence_score,
               ra.refund_calculation_method
        FROM refund_analysis ra
        JOIN invoice_line_items li ON ra.line_item_id = li.id
        JOIN invoices i ON ra.invoice_id = i.id
        WHERE i.client_id = ? AND ra.potentially_eligible = 1 AND ra.confidence_score > 80
    """, (client_id,))

    for i, row in enumerate(cursor.fetchall(), 2):
        for j, val in enumerate(row, 1):
            ws3.cell(i, j, val)

    xl.format_as_table(ws3)

    # Save file
    project_root = Path(__file__).parent.parent
    output_dir = project_root / "outputs" / "analysis"
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{client_name.replace(' ', '_')}_Internal_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)
    conn.close()

    print(f"✅ Internal analysis workbook generated: {filepath}")
    return str(filepath)

def main():
    parser = argparse.ArgumentParser(description="Generate internal analysis workbook")
    parser.add_argument('--client_id', required=True, type=int)
    args = parser.parse_args()

    generate_internal_analysis_workbook(args.client_id)
    return 0

if __name__ == "__main__":
    sys.exit(main())
