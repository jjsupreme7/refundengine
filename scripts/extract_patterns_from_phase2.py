"""
Extract patterns from Phase 2 Master Refunds Excel file and update pattern JSON files
WITH QUALITY FILTERING - No junk data (PDFs, numeric codes, etc.)
"""
import openpyxl
import json
from collections import defaultdict, Counter
from pathlib import Path
import sys

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from pattern_filters import (
    is_valid_refund_basis,
    is_valid_keyword,
    is_valid_category_value,
    is_text_gl_description,
    is_meaningful_material_code,
    is_meaningful_location,
    clean_wac_citations
)

# File paths
EXCEL_FILE = r'C:\Users\jacob\Desktop\Denodo&UseTax\Phase 2 Master Refunds_6-15-25.xlsx'
OUTPUT_DIR = Path(r'c:\Users\jacob\refundengine\refundengine\extracted_patterns')

def load_excel_data(file_path):
    """Load and parse the Excel workbook"""
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")

        # Find header row
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if any(cell and 'Vendor Name' in str(cell) for cell in row):
                header_row = i
                headers = [cell if cell else f'Column_{idx}' for idx, cell in enumerate(row)]
                print(f"Found headers in row {header_row}: {len(headers)} columns")
                break

        if not header_row:
            print(f"No header row found in {sheet_name}, skipping...")
            continue

        # Extract data
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            row_dict = dict(zip(headers, row))
            if row_dict.get('Vendor Name'):  # Only include rows with vendor name
                data.append(row_dict)

    print(f"\nTotal records loaded: {len(data)}")
    return data

def extract_vendor_patterns(data):
    """Extract vendor-specific patterns"""
    vendor_stats = defaultdict(lambda: {
        'transaction_count': 0,
        'refund_bases': Counter(),
        'tax_categories': Counter(),
        'product_types': Counter(),
        'description_keywords': Counter(),
        'allocation_methods': Counter()
    })

    for record in data:
        vendor_name = record.get('Vendor Name', '').strip()
        if not vendor_name:
            continue

        stats = vendor_stats[vendor_name]
        stats['transaction_count'] += 1

        # Refund basis (FILTERED - no PDFs, no invoice numbers)
        refund_basis = record.get('Refund Basis') or record.get('Basis for Refund')
        if is_valid_refund_basis(refund_basis):
            stats['refund_bases'][str(refund_basis).strip()] += 1

        # Tax category (FILTERED - no numbers, no headers)
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            stats['tax_categories'][str(tax_cat).strip()] += 1

        # Product type (FILTERED - no numbers, no headers)
        product_type = record.get('Product Type')
        if is_valid_category_value(product_type):
            stats['product_types'][str(product_type).strip()] += 1

        # Description keywords from Column AG (Description field)
        # FILTERED - no numbers, must have letters, no stopwords
        description = record.get('Description')
        if description and str(description).strip() not in ['', 'nan', 'None', 'See SOW', 'See Excel']:
            words = str(description).split()
            for word in words:
                if is_valid_keyword(word):
                    stats['description_keywords'][word] += 1

        # Allocation method
        allocation = record.get('Allocation Methodology')
        if allocation and str(allocation).strip() not in ['', 'nan', 'None']:
            stats['allocation_methods'][str(allocation).strip()] += 1

    # Convert to vendor patterns format (GL accounts removed per user request)
    vendor_patterns = []
    for vendor_name, stats in sorted(vendor_stats.items()):
        # Get most common values
        top_refund_basis = stats['refund_bases'].most_common(1)[0][0] if stats['refund_bases'] else 'Unknown'
        top_tax_category = stats['tax_categories'].most_common(1)[0][0] if stats['tax_categories'] else 'Unknown'
        top_product_type = stats['product_types'].most_common(1)[0][0] if stats['product_types'] else 'Unknown'
        top_keywords = [word for word, _ in stats['description_keywords'].most_common(10)]

        pattern = {
            'vendor_name': vendor_name,
            'product_type': top_product_type,
            'historical_sample_count': stats['transaction_count'],
            'historical_success_rate': 1.0,  # Assuming all Phase 2 records were successful
            'typical_refund_basis': top_refund_basis,
            'common_tax_categories': [cat for cat, _ in stats['tax_categories'].most_common(3)],
            'common_refund_bases': [rb for rb, _ in stats['refund_bases'].most_common(3)],
            'common_allocation_methods': [am for am, _ in stats['allocation_methods'].most_common(3)],
            'description_keywords': top_keywords[:5] if top_keywords else []
            # GL accounts removed per user request - "I don't care about gl accounts for now"
        }
        vendor_patterns.append(pattern)

    print(f"\nExtracted patterns for {len(vendor_patterns)} vendors")
    return vendor_patterns

def extract_keyword_patterns(data):
    """Extract keyword patterns across all categories"""
    keywords = {
        'tax_categories': Counter(),
        'product_types': Counter(),
        'delivery_methods': Counter(),
        'refund_basis_terms': Counter(),
        'allocation_methods': Counter(),
        'material_group_codes': Counter(),
        'location_keywords': Counter()
    }

    for record in data:
        # Tax categories (FILTERED - no numbers, no headers)
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            keywords['tax_categories'][str(tax_cat).strip()] += 1

        # Product types (FILTERED - no numbers, no headers)
        product_type = record.get('Product Type')
        if is_valid_category_value(product_type):
            keywords['product_types'][str(product_type).strip()] += 1

        # Delivery methods (FILTERED - no numbers, no headers)
        delivery = record.get('Delivery Method')
        if is_valid_category_value(delivery):
            keywords['delivery_methods'][str(delivery).strip()] += 1

        # Refund basis (FILTERED - no PDFs, no invoice numbers)
        refund_basis = record.get('Refund Basis') or record.get('Basis for Refund')
        if is_valid_refund_basis(refund_basis):
            keywords['refund_basis_terms'][str(refund_basis).strip()] += 1

        # Allocation methods (keep text descriptions)
        allocation = record.get('Allocation Methodology')
        if is_valid_category_value(allocation):
            keywords['allocation_methods'][str(allocation).strip()] += 1

        # Material group codes (FILTERED - only codes with letters, or skip if all numeric)
        material = record.get('PO Material Group')
        if is_meaningful_material_code(material):
            keywords['material_group_codes'][str(material).strip()] += 1

        # Location keywords (FILTERED - meaningful locations only)
        location = record.get('Initial Delivery Location')
        if location and str(location).strip() not in ['', 'nan', 'None']:
            # Split on common delimiters
            locs = str(location).replace(',', ' ').replace(';', ' ').split()
            for loc in locs:
                if is_meaningful_location(loc):
                    keywords['location_keywords'][loc] += 1

    # Format as structured JSON
    keyword_patterns = []

    for category, counter in keywords.items():
        if counter:
            category_data = {
                'category': category,
                'description': get_category_description(category),
                'keywords': [
                    {'term': term, 'count': count}
                    for term, count in counter.most_common(50)  # Top 50 per category
                ]
            }
            keyword_patterns.append(category_data)

    print(f"\nExtracted {len(keyword_patterns)} keyword categories")
    return keyword_patterns

def get_category_description(category):
    """Get description for keyword category"""
    descriptions = {
        'tax_categories': 'Primary tax classification categories',
        'product_types': 'Types of products and services',
        'delivery_methods': 'How software/services are delivered',
        'refund_basis_terms': 'Reasons for refund claims',
        'allocation_methods': 'Methods used to allocate tax across jurisdictions',
        'material_group_codes': 'PO material classification codes',
        'location_keywords': 'Common locations referenced in transactions'
    }
    return descriptions.get(category, '')

def extract_refund_basis_patterns(data):
    """Extract refund basis patterns - FILTERED (no PDFs, no invoice numbers)
    Returns ALL vendors per refund basis (not just 5 examples) per user requirement
    """
    refund_basis_stats = Counter()
    refund_basis_vendors = defaultdict(set)

    total_records = len(data)

    for record in data:
        refund_basis = record.get('Refund Basis') or record.get('Basis for Refund')
        # Apply filter - no PDFs, no invoice numbers, must be text
        if is_valid_refund_basis(refund_basis):
            rb = str(refund_basis).strip()
            refund_basis_stats[rb] += 1

            vendor = record.get('Vendor Name', '').strip()
            if vendor:
                refund_basis_vendors[rb].add(vendor)

    patterns = []
    for refund_basis, count in refund_basis_stats.most_common():
        pattern = {
            'refund_basis': refund_basis,
            'usage_count': count,
            'percentage': round(count / total_records * 100, 1),
            'all_vendors': sorted(list(refund_basis_vendors[refund_basis])),  # ALL vendors, sorted alphabetically
            'vendor_count': len(refund_basis_vendors[refund_basis])
        }
        patterns.append(pattern)

    print(f"\nExtracted {len(patterns)} refund basis patterns")
    return patterns

def extract_citation_patterns(data):
    """
    Extract legal citation patterns - ONLY IF REAL EXAMPLES EXIST
    Focus on TEXT descriptions, NO numeric codes per user requirement
    """
    import re

    wac_examples = []
    wac_pattern = re.compile(r'WAC\s+\d{3}-\d{2}-\d{3,4}[A-Z]?', re.IGNORECASE)
    gl_descriptions = Counter()  # TEXT descriptions only, not numeric codes

    for record in data:
        # Look for WAC patterns in all text fields - must match actual pattern
        for key, value in record.items():
            if value and isinstance(value, str):
                if wac_pattern.search(value):
                    wac_examples.append(value)

        # GL Account - TEXT descriptions only (user wants NO numeric codes)
        gl_account = record.get('GL Account') or record.get('G/L Account')
        if is_text_gl_description(gl_account):
            gl_descriptions[str(gl_account).strip()] += 1

    # Clean WAC citations - only keep actual matches
    wac_examples = clean_wac_citations(wac_examples)

    patterns = []

    # Only add WAC if we found real examples
    if wac_examples:
        patterns.append({
            'citation_type': 'WAC',
            'pattern': r'WAC \d{3}-\d{2}-\d{3,4}[A-Z]?',
            'examples': wac_examples[:10],
            'description': 'Washington Administrative Code tax regulations'
        })

    # Only add GL descriptions if we have text-based ones (no numeric codes)
    if gl_descriptions:
        patterns.append({
            'citation_type': 'GL_Account_Descriptions',
            'description': 'General Ledger account TEXT descriptions (numeric codes excluded per user requirement)',
            'descriptions': [
                {'description': desc, 'frequency': freq}
                for desc, freq in gl_descriptions.most_common(20)
            ]
        })

    # RCW removed - no examples found, user said don't force patterns

    print(f"\nExtracted citation patterns: {len(wac_examples)} WAC examples, {len(gl_descriptions)} GL descriptions (TEXT only)")
    return patterns

def merge_with_existing_vendors(new_patterns, existing_file):
    """Merge new vendor patterns with existing ones"""
    if not existing_file.exists():
        print("No existing vendor patterns file found, using new data only")
        return new_patterns

    with open(existing_file, 'r', encoding='utf-8') as f:
        existing_patterns = json.load(f)

    # Create lookup by vendor name
    existing_by_name = {p['vendor_name']: p for p in existing_patterns}

    merged = []
    for new_pattern in new_patterns:
        vendor_name = new_pattern['vendor_name']

        if vendor_name in existing_by_name:
            # Merge: add transaction counts and update fields
            existing = existing_by_name[vendor_name]
            existing['historical_sample_count'] += new_pattern['historical_sample_count']

            # Update with newer data
            if new_pattern.get('typical_refund_basis') != 'Unknown':
                existing['typical_refund_basis'] = new_pattern['typical_refund_basis']

            # Add new fields
            existing['common_tax_categories'] = new_pattern.get('common_tax_categories', [])
            existing['common_refund_bases'] = new_pattern.get('common_refund_bases', [])
            existing['common_allocation_methods'] = new_pattern.get('common_allocation_methods', [])

            merged.append(existing)
            del existing_by_name[vendor_name]
        else:
            # New vendor
            merged.append(new_pattern)

    # Add remaining existing vendors that weren't in new data
    merged.extend(existing_by_name.values())

    # Sort by transaction count descending
    merged.sort(key=lambda x: x['historical_sample_count'], reverse=True)

    print(f"\nMerged {len(new_patterns)} new vendors with {len(existing_patterns)} existing vendors")
    print(f"Total: {len(merged)} vendors ({len([p for p in merged if p in new_patterns])} new, {len(existing_by_name)} unchanged)")

    return merged

def main():
    """Main execution"""
    print("=" * 80)
    print("EXTRACTING PATTERNS FROM PHASE 2 MASTER REFUNDS")
    print("=" * 80)

    # Load data
    data = load_excel_data(EXCEL_FILE)

    if not data:
        print("ERROR: No data loaded from Excel file")
        return

    # Extract patterns
    print("\n" + "=" * 80)
    print("EXTRACTING VENDOR PATTERNS")
    print("=" * 80)
    vendor_patterns = extract_vendor_patterns(data)

    # Merge with existing
    vendor_file = OUTPUT_DIR / 'vendor_patterns.json'
    merged_vendors = merge_with_existing_vendors(vendor_patterns, vendor_file)

    print("\n" + "=" * 80)
    print("EXTRACTING KEYWORD PATTERNS")
    print("=" * 80)
    keyword_patterns = extract_keyword_patterns(data)

    print("\n" + "=" * 80)
    print("EXTRACTING REFUND BASIS PATTERNS")
    print("=" * 80)
    refund_basis_patterns = extract_refund_basis_patterns(data)

    # Citation patterns removed per user request - "I don't need the citations pattern"

    # Save all patterns
    print("\n" + "=" * 80)
    print("SAVING PATTERN FILES")
    print("=" * 80)

    # Save vendor patterns
    vendor_output = OUTPUT_DIR / 'vendor_patterns.json'
    with open(vendor_output, 'w', encoding='utf-8') as f:
        json.dump(merged_vendors, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(merged_vendors)} vendor patterns to {vendor_output}")

    # Save keyword patterns
    keyword_output = OUTPUT_DIR / 'keyword_patterns.json'
    with open(keyword_output, 'w', encoding='utf-8') as f:
        json.dump(keyword_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(keyword_patterns)} keyword categories to {keyword_output}")

    # Save refund basis patterns
    refund_basis_output = OUTPUT_DIR / 'refund_basis_patterns.json'
    with open(refund_basis_output, 'w', encoding='utf-8') as f:
        json.dump(refund_basis_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(refund_basis_patterns)} refund basis patterns to {refund_basis_output}")

    print("\n" + "=" * 80)
    print("COMPLETE!")
    print("=" * 80)
    print(f"\nTotal records processed: {len(data)}")
    print(f"Vendors: {len(merged_vendors)}")
    print(f"Keyword categories: {len(keyword_patterns)}")
    print(f"Refund basis patterns: {len(refund_basis_patterns)}")

if __name__ == '__main__':
    main()
