#!/usr/bin/env python3
"""
Excel Output Formatting Utility

Applies consistent formatting to analyzed tax output files:
- Color-coded headers (blue=input, green=analysis, yellow=review)
- Number formatting (commas for currency, percentages for rates)
- Auto-fit column widths

Usage:
    python scripts/format_excel_output.py <filepath>
    python scripts/format_excel_output.py  # formats all output files
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Header colors by column group
HEADER_INPUT = PatternFill("solid", fgColor="B4C6E7")      # Light blue - input data
HEADER_ANALYSIS = PatternFill("solid", fgColor="C6EFCE")   # Light green - AI analysis
HEADER_REVIEW = PatternFill("solid", fgColor="FFEB9C")     # Light yellow - human review

# Row highlight colors by decision
ROW_OVERPAID = PatternFill("solid", fgColor="E2EFDA")      # Light green
ROW_REVIEW = PatternFill("solid", fgColor="FFF2CC")        # Light yellow
ROW_NO_TAX = PatternFill("solid", fgColor="D9D9D9")        # Light gray

# Column groups (based on CLAUDE.md column order)
# Sales Tax 2024: Input (1-13), Analysis (14-29), Review (30-36)
# Use Tax: Input (1-8), Analysis (9-24), Review (25-31)
INPUT_COLUMNS_SALES = 13
ANALYSIS_COLUMNS_SALES = 29
INPUT_COLUMNS_USE = 8
ANALYSIS_COLUMNS_USE = 24

# Columns that need number formatting
CURRENCY_COLUMNS = ['hwbas_tax_base_lc', 'hwste_tax_amount_lc', 'Estimated_Refund', 'Tax Remitted']
RATE_COLUMNS = ['rate']


def apply_formatting(filepath: str, tax_type: str = "auto") -> None:
    """
    Apply formatting to an analyzed Excel output file.

    Args:
        filepath: Path to the Excel file
        tax_type: "sales", "use", or "auto" (detect from filename)
    """
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"Error: File not found: {filepath}")
        return

    print(f"Formatting: {filepath.name}")

    # Auto-detect tax type from filename
    if tax_type == "auto":
        name_lower = filepath.name.lower()
        if "use" in name_lower:
            tax_type = "use"
        else:
            tax_type = "sales"

    # Set column boundaries based on tax type
    if tax_type == "use":
        input_end = INPUT_COLUMNS_USE
        analysis_end = ANALYSIS_COLUMNS_USE
    else:
        input_end = INPUT_COLUMNS_SALES
        analysis_end = ANALYSIS_COLUMNS_SALES

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        print(f"Error opening file: {e}")
        return

    ws = wb.active

    # Build column name to index mapping
    col_names = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            col_names[cell.value] = col_idx

    # === HEADER FORMATTING ===
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF" if col_idx <= input_end else "000000")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

        if col_idx <= input_end:
            cell.fill = HEADER_INPUT
        elif col_idx <= analysis_end:
            cell.fill = HEADER_ANALYSIS
        else:
            cell.fill = HEADER_REVIEW

    # === NUMBER FORMATTING ===
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name = ws.cell(1, cell.column).value

            # Currency formatting (commas, 2 decimals, no $)
            if col_name in CURRENCY_COLUMNS:
                cell.number_format = '#,##0.00'

            # Percentage formatting
            elif col_name in RATE_COLUMNS:
                cell.number_format = '0.0%'

    # === ROW HIGHLIGHTING BY DECISION ===
    decision_col = col_names.get('Final_Decision')
    if decision_col:
        for row_idx in range(2, ws.max_row + 1):
            decision = ws.cell(row_idx, decision_col).value
            if decision == 'OVERPAID':
                fill = ROW_OVERPAID
            elif decision == 'REVIEW':
                fill = ROW_REVIEW
            elif decision == 'NO_TAX_CHARGED':
                fill = ROW_NO_TAX
            else:
                continue

            # Apply to non-header cells in the row
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                # Don't override if cell is empty in review section
                if col_idx > analysis_end and not cell.value:
                    continue
                cell.fill = fill

    # === COLUMN WIDTHS ===
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = ws.cell(1, col[0].column).value

        # Calculate max content length
        max_length = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value or ''))
                max_length = max(max_length, cell_len)
            except:
                pass

        # Set width with limits
        if col_name == 'AI_Reasoning':
            width = 80  # Wide for detailed reasoning
        elif col_name in ['Explanation', 'Follow_Up_Questions']:
            width = 60
        elif col_name in CURRENCY_COLUMNS:
            width = 15  # Standard width for currency
        else:
            width = min(max_length + 2, 40)  # Cap at 40

        ws.column_dimensions[col_letter].width = max(width, 10)  # Minimum 10

    # === FREEZE HEADER ROW ===
    ws.freeze_panes = 'A2'

    # Save the file
    try:
        wb.save(filepath)
        print(f"  Formatted successfully!")
    except PermissionError:
        backup_path = filepath.with_suffix('.formatted.xlsx')
        wb.save(backup_path)
        print(f"  Original file locked, saved to: {backup_path}")


def format_all_output_files():
    """Format all standard output files in the Analyzed_Output directory."""
    output_dir = Path.home() / "Desktop/Files/Analyzed_Output"

    files_to_format = [
        "Final 2024 Denodo Review - Analyzed.xlsx",
        "Phase_3_2023_Use Tax - Analyzed.xlsx",
        "Phase_3_2024_Use Tax - Analyzed.xlsx"
    ]

    for filename in files_to_format:
        filepath = output_dir / filename
        if filepath.exists():
            apply_formatting(str(filepath))
        else:
            print(f"Skipping (not found): {filename}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        # Format specific file
        apply_formatting(sys.argv[1])
    else:
        # Format all standard output files
        format_all_output_files()
