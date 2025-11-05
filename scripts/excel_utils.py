"""
Excel Utility Functions for Washington State Tax Refund Engine

Reusable Excel formatting and styling functions.
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_currency(worksheet, cell_range):
    """Apply currency formatting to cells."""
    for row in worksheet[cell_range]:
        for cell in row if isinstance(row, tuple) else [row]:
            cell.number_format = '$#,##0.00'

def format_date(worksheet, cell_range):
    """Apply date formatting to cells."""
    for row in worksheet[cell_range]:
        for cell in row if isinstance(row, tuple) else [row]:
            cell.number_format = 'MM/DD/YYYY'

def format_percentage(worksheet, cell_range):
    """Apply percentage formatting to cells."""
    for row in worksheet[cell_range]:
        for cell in row if isinstance(row, tuple) else [row]:
            cell.number_format = '0%'

def apply_alternating_rows(worksheet, start_row, end_row, color1='FFFFFF', color2='F2F2F2'):
    """Apply alternating row colors."""
    for row_num in range(start_row, end_row + 1):
        fill_color = color1 if row_num % 2 == 0 else color2
        for cell in worksheet[row_num]:
            if cell.fill.patternType is None or cell.fill.fgColor.rgb == '00000000':
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

def auto_adjust_columns(worksheet, min_width=10, max_width=80):
    """Auto-adjust column widths based on content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def add_header_row(worksheet, headers, row=1, bold=True, bg_color='4472C4', font_color='FFFFFF'):
    """Add formatted header row."""
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=bold, color=font_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def freeze_top_row(worksheet):
    """Freeze the top row."""
    worksheet.freeze_panes = 'A2'

def enable_auto_filter(worksheet, end_column='Z'):
    """Enable auto-filter on headers."""
    worksheet.auto_filter.ref = f'A1:{end_column}1'

def add_conditional_formatting_confidence(worksheet, column, start_row, end_row):
    """Add conditional formatting for confidence scores."""
    from openpyxl.formatting.rule import ColorScaleRule

    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=50, mid_color='FFEB84',
        end_type='num', end_value=100, end_color='63BE7B'
    )

    worksheet.conditional_formatting.add(
        f'{get_column_letter(column)}{start_row}:{get_column_letter(column)}{end_row}',
        rule
    )

def add_borders(worksheet, cell_range, style='thin'):
    """Add borders to cells."""
    border = Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style)
    )

    for row in worksheet[cell_range]:
        for cell in row if isinstance(row, tuple) else [row]:
            cell.border = border

def merge_and_center(worksheet, start_cell, end_cell, value, font_size=14, bold=True):
    """Merge cells and center text."""
    worksheet.merge_cells(f'{start_cell}:{end_cell}')
    cell = worksheet[start_cell]
    cell.value = value
    cell.font = Font(size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def add_title(worksheet, title, row=1, start_col=1, end_col=10):
    """Add a title spanning multiple columns."""
    start_cell = f'{get_column_letter(start_col)}{row}'
    end_cell = f'{get_column_letter(end_col)}{row}'
    merge_and_center(worksheet, start_cell, end_cell, title, font_size=16, bold=True)
    worksheet.row_dimensions[row].height = 30

def format_as_table(worksheet, start_row=1, end_row=None, end_col=None):
    """Format range as a professional table."""
    if end_row is None:
        end_row = worksheet.max_row
    if end_col is None:
        end_col = worksheet.max_column

    # Add borders
    add_borders(worksheet, f'A{start_row}:{get_column_letter(end_col)}{end_row}')

    # Format header
    for col in range(1, end_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Alternating rows
    if end_row > start_row:
        apply_alternating_rows(worksheet, start_row + 1, end_row)

    # Auto-adjust columns
    auto_adjust_columns(worksheet)

    # Freeze header
    worksheet.freeze_panes = f'A{start_row + 1}'
