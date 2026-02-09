#!/usr/bin/env python3
"""Benchmark refund engine methodology + allocation against Phase 2 human decisions.

Usage:
    python scripts/benchmark_against_phase2.py [path_to_master_file]

Reads the Phase 2 Master Refunds file, compares methodology assignments and
allocation percentages against vendor_profiles.json and allocation_percentages.json.
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROFILES_PATH = PROJECT_ROOT / "config" / "vendor_profiles.json"
ALLOC_PATH = PROJECT_ROOT / "config" / "allocation_percentages.json"


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path) as f:
        return json.load(f)


def run_benchmark(master_path: str) -> None:
    print(f"Reading {master_path}...")
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb["Refund Summary"]

    profiles = _load_json(PROFILES_PATH).get("vendors", {})
    alloc_config = _load_json(ALLOC_PATH)

    # Collect per-vendor data from master file
    # Columns: A=1 Vendor, T=20 Sales/Use, V=22 Tax Paid, X=24 Non-Taxable%, Y=25 Methodology
    vendor_data: dict[str, dict] = defaultdict(lambda: {
        "rows": 0, "benchmarkable": 0,
        "methodologies": defaultdict(int),
        "phase2_pcts": [], "system_pcts": [], "profile_pcts": [],
    })

    total_rows = 0
    benchmarkable = 0

    for r in range(2, ws.max_row + 1):
        vendor = ws.cell(r, 1).value
        if not vendor:
            continue
        vendor = str(vendor).strip().upper()
        total_rows += 1

        methodology = ws.cell(r, 25).value
        tax_paid = ws.cell(r, 22).value
        nontax_pct = ws.cell(r, 24).value

        vd = vendor_data[vendor]
        vd["rows"] += 1
        if methodology:
            vd["methodologies"][str(methodology)] += 1

        # Need all three for benchmarking
        if not (methodology and tax_paid is not None and nontax_pct is not None):
            continue
        try:
            tax_paid_f = float(tax_paid)
            nontax_pct_f = float(nontax_pct)
            if not (0.0 <= nontax_pct_f <= 1.0 and tax_paid_f > 0):
                continue
        except (TypeError, ValueError):
            continue

        benchmarkable += 1
        vd["benchmarkable"] += 1
        vd["phase2_pcts"].append(nontax_pct_f)

        meth_str = str(methodology)
        alloc_entry = alloc_config.get(meth_str, {})
        system_pct = alloc_entry.get("allocation_pct") if isinstance(alloc_entry, dict) else None
        if system_pct is not None:
            vd["system_pcts"].append(float(system_pct))

        # Check vendor profile for methodology-specific avg_pct
        profile = profiles.get(vendor, {})
        mix = profile.get("methodology_mix", {})
        meth_info = mix.get(meth_str, {})
        prof_pct = meth_info.get("avg_pct")
        if prof_pct is not None:
            vd["profile_pcts"].append(float(prof_pct))

    # Report
    print(f"\n{'='*70}")
    print("BENCHMARK REPORT: Phase 2 vs Refund Engine")
    print(f"{'='*70}")
    print(f"Total rows: {total_rows}")
    print(f"Benchmarkable rows (Tax Paid + Methodology + Non-Taxable%): {benchmarkable}")
    print(f"Vendors in profiles: {len(profiles)}")
    print()

    meth_match = 0
    meth_total = 0
    system_better = 0
    profile_better = 0
    comparable = 0
    total_system_var = 0.0
    total_profile_var = 0.0
    system_var_count = 0
    profile_var_count = 0

    # Top vendors by row count
    sorted_vendors = sorted(vendor_data.items(), key=lambda x: -x[1]["rows"])

    print("TOP 15 VENDORS:")
    print("-" * 70)
    for vendor, vd in sorted_vendors[:15]:
        meths = vd["methodologies"]
        dominant_meth = max(meths, key=meths.get) if meths else "N/A"
        dominant_count = meths.get(dominant_meth, 0)

        profile = profiles.get(vendor, {})
        profile_dominant = profile.get("dominant_methodology", {}).get("value", "N/A")

        matched = dominant_meth == profile_dominant
        match_mark = "Y" if matched else "X"

        avg_p2 = sum(vd["phase2_pcts"]) / len(vd["phase2_pcts"]) if vd["phase2_pcts"] else None
        avg_sys = sum(vd["system_pcts"]) / len(vd["system_pcts"]) if vd["system_pcts"] else None
        avg_prof = sum(vd["profile_pcts"]) / len(vd["profile_pcts"]) if vd["profile_pcts"] else None

        print(f"\n  {vendor} ({vd['rows']} rows, {vd['benchmarkable']} benchmarkable)")
        print(f"    Phase 2 dominant: {dominant_meth} ({dominant_count} rows)")
        print(f"    Profile dominant: {profile_dominant} [{match_mark}]")
        if avg_p2 is not None:
            print(f"    Phase 2 avg alloc: {avg_p2:.1%}")
        if avg_sys is not None:
            var = avg_sys - (avg_p2 or 0)
            print(f"    System global pct: {avg_sys:.1%} (var: {var:+.1%}pp)")
        if avg_prof is not None:
            var = avg_prof - (avg_p2 or 0)
            print(f"    Profile avg pct:  {avg_prof:.1%} (var: {var:+.1%}pp)")

    # Aggregate stats
    for vendor, vd in vendor_data.items():
        meths = vd["methodologies"]
        if not meths:
            continue
        dominant_meth = max(meths, key=meths.get)
        profile = profiles.get(vendor, {})
        profile_dominant = profile.get("dominant_methodology", {}).get("value")
        if profile_dominant:
            meth_total += 1
            if dominant_meth == profile_dominant:
                meth_match += 1

        avg_p2 = sum(vd["phase2_pcts"]) / len(vd["phase2_pcts"]) if vd["phase2_pcts"] else None
        avg_sys = sum(vd["system_pcts"]) / len(vd["system_pcts"]) if vd["system_pcts"] else None
        avg_prof = sum(vd["profile_pcts"]) / len(vd["profile_pcts"]) if vd["profile_pcts"] else None

        if avg_p2 is not None and avg_sys is not None:
            total_system_var += abs(avg_sys - avg_p2)
            system_var_count += 1
        if avg_p2 is not None and avg_prof is not None:
            total_profile_var += abs(avg_prof - avg_p2)
            profile_var_count += 1

        if avg_p2 is not None and avg_sys is not None and avg_prof is not None:
            comparable += 1
            sys_err = abs(avg_sys - avg_p2)
            prof_err = abs(avg_prof - avg_p2)
            if prof_err < sys_err:
                profile_better += 1
            elif sys_err < prof_err:
                system_better += 1

    print(f"\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}")
    print(f"  Methodology match rate: {meth_match}/{meth_total} ({meth_match/meth_total:.0%})" if meth_total else "  Methodology match rate: N/A")
    if system_var_count:
        print(f"  Avg system allocation variance: {total_system_var/system_var_count:.1%}pp")
    if profile_var_count:
        print(f"  Avg profile allocation variance: {total_profile_var/profile_var_count:.1%}pp")
    if comparable:
        print(f"  Profile more accurate than global: {profile_better}/{comparable} ({profile_better/comparable:.0%})")
        print(f"  Global more accurate than profile: {system_better}/{comparable} ({system_better/comparable:.0%})")


if __name__ == "__main__":
    default_master = str(
        Path.home() / "Downloads" / "Phase 2 Master Refunds Jun 15 2025.xlsx"
    )
    master = sys.argv[1] if len(sys.argv) > 1 else default_master
    run_benchmark(master)
