#!/usr/bin/env python3
"""
Client Report Generator for Washington State Tax Refund Engine

Creates professional 4-sheet client-facing Excel report.

Usage:
    python scripts/generate_client_report.py --client_id 1
"""

import argparse
import sys
import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
import excel_utils as xl

def get_database_connection():
    """Get database connection."""
    project_root = Path(__file__).parent.parent
    db_path = project_root / "database" / "refund_engine.db"
    return sqlite3.connect(str(db_path))

def generate_client_refund_report(client_id):
    """Generate comprehensive client refund report."""
    conn = get_database_connection()
    cursor = conn.cursor()

    # Get client info
    cursor.execute("SELECT client_name FROM clients WHERE id = ?", (client_id,))
    client = cursor.fetchone()
    if not client:
        print(f"❌ Client {client_id} not found")
        return None

    client_name = client[0]

    print(f"\n📊 Generating client report for: {client_name}")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # SHEET 1: Executive Summary
    ws1 = wb.create_sheet("Executive Summary")

    xl.add_title(ws1, f"{client_name} - Tax Refund Analysis", row=1, end_col=5)
    ws1.cell(2, 1, f"Report Generated: {datetime.now().strftime('%B %d, %Y')}")

    # Get summary stats
    cursor.execute("""
        SELECT
            COUNT(DISTINCT ra.invoice_id) as invoice_count,
            SUM(CASE WHEN ra.potentially_eligible = 1 THEN ra.estimated_refund_amount ELSE 0 END) as total_refund,
            AVG(ra.confidence_score) as avg_confidence,
            COUNT(CASE WHEN ra.potentially_eligible = 1 THEN 1 END) as eligible_items
        FROM refund_analysis ra
        JOIN invoices i ON ra.invoice_id = i.id
        WHERE i.client_id = ?
    """, (client_id,))

    stats = cursor.fetchone()
    invoice_count, total_refund, avg_confidence, eligible_items = stats

    # Summary metrics
    ws1.cell(4, 1, "KEY METRICS").font = Font(bold=True, size=12)
    metrics = [
        ("Total Potential Refund:", f"${total_refund or 0:.2f}"),
        ("Invoices Analyzed:", invoice_count or 0),
        ("Eligible Line Items:", eligible_items or 0),
        ("Average Confidence:", f"{avg_confidence or 0:.1f}%")
    ]

    row = 5
    for label, value in metrics:
        ws1.cell(row, 1, label).font = Font(bold=True)
        ws1.cell(row, 2, value)
        row += 1

    xl.auto_adjust_columns(ws1)

    # SHEET 2: Refund Details
    ws2 = wb.create_sheet("Refund Details")

    headers = ["Invoice #", "Date", "Vendor", "Description", "Tax Paid", "Refund", "Exemption", "Citation", "Confidence"]
    xl.add_header_row(ws2, headers)

    cursor.execute("""
        SELECT
            i.invoice_number,
            i.invoice_date,
            i.vendor_name,
            li.item_description,
            li.sales_tax_on_line,
            ra.estimated_refund_amount,
            ra.refund_calculation_method,
            ra.criteria_matching_json,
            ra.confidence_score
        FROM refund_analysis ra
        JOIN invoices i ON ra.invoice_id = i.id
        JOIN invoice_line_items li ON ra.line_item_id = li.id
        WHERE i.client_id = ? AND ra.potentially_eligible = 1
        ORDER BY ra.estimated_refund_amount DESC
    """, (client_id,))

    row = 2
    for rec in cursor.fetchall():
        inv_num, inv_date, vendor, desc, tax, refund, method, criteria, conf = rec
        ws2.cell(row, 1, inv_num)
        ws2.cell(row, 2, inv_date)
        ws2.cell(row, 3, vendor)
        ws2.cell(row, 4, desc[:50])
        ws2.cell(row, 5, tax or 0)
        ws2.cell(row, 6, refund or 0)
        ws2.cell(row, 7, method[:30] if method else "")
        ws2.cell(row, 8, "See analysis")
        ws2.cell(row, 9, conf or 0)
        row += 1

    xl.format_currency(ws2, f'E2:F{row-1}')
    xl.format_as_table(ws2, end_row=row-1, end_col=9)

    # SHEET 3: Documentation Checklist
    ws3 = wb.create_sheet("Documentation Checklist")
    xl.add_header_row(ws3, ["Document Type", "Status", "Notes"])

    docs = [
        ("Original invoices", "Required", "All invoices with eligible items"),
        ("Proof of tax paid", "Required", "Payment confirmation"),
        ("Supporting documentation", "As needed", "Per exemption type"),
    ]

    for i, (doc_type, status, notes) in enumerate(docs, 2):
        ws3.cell(i, 1, doc_type)
        ws3.cell(i, 2, status)
        ws3.cell(i, 3, notes)

    xl.auto_adjust_columns(ws3)

    # SHEET 4: Legal References
    ws4 = wb.create_sheet("Legal References")
    xl.add_header_row(ws4, ["Citation", "Description"])

    ws4.cell(2, 1, "RCW 82.08.02565")
    ws4.cell(2, 2, "Manufacturing equipment exemption")

    xl.auto_adjust_columns(ws4)

    # Save file
    project_root = Path(__file__).parent.parent
    output_dir = project_root / "outputs" / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{client_name.replace(' ', '_')}_Refund_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)
    conn.close()

    print(f"✅ Client report generated: {filepath}")
    return str(filepath)

def main():
    """Main CLI function."""
    parser = argparse.ArgumentParser(description="Generate client refund report")
    parser.add_argument('--client_id', required=True, type=int, help="Client ID")
    args = parser.parse_args()

    generate_client_refund_report(args.client_id)
    return 0

if __name__ == "__main__":
    sys.exit(main())
