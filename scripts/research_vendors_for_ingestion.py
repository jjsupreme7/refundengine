#!/usr/bin/env python3
"""
Research vendors and create Vendor_Background.xlsx for review before ingestion

This script takes a list of vendor names, researches them using AI,
and creates an Excel file ready for review and import.
"""

import os
import sys
from pathlib import Path
from openai import OpenAI
from dotenv import load_dotenv
import pandas as pd
from tqdm import tqdm
import json

# Load environment
load_dotenv()
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


def research_vendor(vendor_name: str) -> dict:
    """
    Research a vendor and extract structured metadata
    """

    prompt = f"""You are a business research assistant. Research the company "{vendor_name}" and provide structured information.

IMPORTANT: Base your response on your knowledge of this company. Provide your best assessment even if you're not 100% certain.

Return a JSON object with the following fields:

{{
  "vendor_name": "Official company name",
  "industry": "Primary industry (e.g., Technology, Professional Services, Manufacturing, etc.)",
  "business_model": "Business model type (e.g., B2B SaaS, B2B Services, Manufacturing, Retail, etc.)",
  "primary_products": ["Product/service 1", "Product/service 2", "Product/service 3"],
  "typical_delivery": "How products/services are delivered (e.g., Cloud-based, On-premise, Hybrid, In-person, Physical goods)",
  "tax_notes": "Tax-relevant notes for Washington state sales/use tax (e.g., 'Digital automated services', 'Professional services - likely exempt', 'Physical goods - taxable', 'SaaS - subject to B&O tax')",
  "confidence_score": 85.0,
  "document_title": "Vendor Background: [Company Name]",
  "document_summary": "Brief 1-2 sentence summary of what this company does"
}}

For tax_notes, consider:
- Digital products and SaaS are generally subject to WA sales tax
- Professional services are often exempt
- Physical goods are taxable
- Manufacturing and construction have special rules

Be specific and accurate. Return ONLY the JSON object, no other text."""

    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are a business research expert. Provide accurate, structured vendor information in JSON format.",
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.3,
            response_format={"type": "json_object"},
        )

        result = json.loads(response.choices[0].message.content)
        return result

    except Exception as e:
        print(f"❌ Error researching {vendor_name}: {e}")
        return None


def create_vendor_background_xlsx(vendors: list, output_path: str):
    """
    Research vendors and create Vendor_Background.xlsx
    """

    print(f"\n🔍 Researching {len(vendors)} vendors...")
    print("This will use AI to gather vendor metadata\n")

    metadata_rows = []

    for vendor_name in tqdm(vendors, desc="Researching vendors"):
        try:
            # Research vendor
            metadata = research_vendor(vendor_name)

            if not metadata:
                print(f"⚠️  Skipping {vendor_name} - research failed")
                continue

            # Generate synthetic filename for manual entry
            clean_vendor_name = (
                metadata.get("vendor_name", vendor_name)
                .replace(" ", "_")
                .replace("/", "_")
                .replace("\\", "_")
            )
            synthetic_filename = f"{clean_vendor_name}.manual"

            # Build row for Excel
            row = {
                "File_Name": f"{vendor_name} (Manual Entry)",
                "File_Path": synthetic_filename,
                "Total_Pages": 0,
                "Document_Type": "vendor_background",
                "Status": "Review",  # User will change to: Approved or Skip
                "document_title": metadata.get(
                    "document_title", f"Vendor Background: {vendor_name}"
                ),
                "document_summary": metadata.get("document_summary", ""),
                "vendor_name": metadata.get("vendor_name", vendor_name),
                "vendor_category": "service_provider",  # Default, user can edit
                "industry": metadata.get("industry", ""),
                "business_model": metadata.get("business_model", ""),
                "primary_products": ", ".join(metadata.get("primary_products", [])),
                "typical_delivery": metadata.get("typical_delivery", ""),
                "tax_notes": metadata.get("tax_notes", ""),
                "document_category": "background",
                "product_types": "",  # User can fill in
                "industries": metadata.get(
                    "industry", ""
                ),  # Use same as industry field
                "keywords": "",  # User can fill in
                "confidence_score": metadata.get("confidence_score", 75.0),
                "AI_Confidence": "Medium",
            }

            metadata_rows.append(row)

        except Exception as e:
            print(f"❌ Error processing {vendor_name}: {e}")
            continue

    # Create DataFrame
    df = pd.DataFrame(metadata_rows)

    # Reorder columns for better Excel layout
    column_order = [
        "File_Name",
        "Status",
        "Document_Type",
        "document_title",
        "vendor_name",
        "vendor_category",
        "industry",
        "business_model",
        "primary_products",
        "typical_delivery",
        "tax_notes",
        "document_category",
        "product_types",
        "industries",
        "keywords",
        "confidence_score",
        "document_summary",
        "AI_Confidence",
        "Total_Pages",
        "File_Path",
    ]

    df = df[column_order]

    # Export to Excel
    df.to_excel(output_path, sheet_name="Metadata", index=False, engine="openpyxl")

    # Format the Excel file
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(output_path)
    ws = wb["Metadata"]

    # Style header row
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)

    print(f"\n{'='*70}")
    print(f"✅ VENDOR RESEARCH COMPLETE")
    print(f"{'='*70}")
    print(f"📊 Researched: {len(metadata_rows)} vendors")
    print(f"📁 Excel file created: {output_path}")
    print(f"\n📋 Next steps:")
    print(f"  1. Open {output_path}")
    print(f"  2. Review AI-suggested metadata for each vendor")
    print(f"  3. Edit any fields you want to change")
    print(f"  4. Update 'Status' column:")
    print(f"     - 'Approved' = Ready to ingest")
    print(f"     - 'Skip' = Don't ingest this vendor")
    print(f"  5. Save the Excel file")
    print(
        f"  6. Run: python core/ingest_documents.py --import-metadata {output_path} --yes"
    )
    print(f"{'='*70}\n")


if __name__ == "__main__":
    # Read vendors from file
    vendors_file = "/tmp/selected_vendors.txt"

    if not Path(vendors_file).exists():
        print(f"❌ Vendors file not found: {vendors_file}")
        sys.exit(1)

    with open(vendors_file, "r") as f:
        vendors = [line.strip() for line in f if line.strip()]

    output_path = "outputs/Vendor_Background.xlsx"
    create_vendor_background_xlsx(vendors, output_path)
