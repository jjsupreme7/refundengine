"""
Extract patterns from Phase 3 2021 Use Tax file - analyst-reviewed records only
Filters for records where 'Final Decision' column is populated
WITH QUALITY FILTERING - No junk data (PDFs, numeric codes, etc.)
"""
import openpyxl
import json
from collections import defaultdict, Counter
from pathlib import Path
import sys

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from pattern_filters import (
    is_valid_refund_basis,
    is_valid_keyword,
    is_valid_category_value
)

# File paths
USE_TAX_FILE = r'C:\Users\jacob\Desktop\Denodo&UseTax\Phase_3_2021_Use Tax_10-17-25.xlsx'
OUTPUT_DIR = Path(r'c:\Users\jacob\refundengine\refundengine\extracted_patterns\use_tax')

def load_use_tax_data(file_path, sheet_name='2021'):
    """Load and parse Use Tax Excel file, filtering for analyst-reviewed records"""
    print(f"Loading Use Tax file: {Path(file_path).name}")
    print(f"Sheet: {sheet_name}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found")
        print(f"Available sheets: {wb.sheetnames}")
        wb.close()
        return []

    ws = wb[sheet_name]

    # Find header row (should be row 1 based on analysis)
    header_row = 1
    headers = [cell.value for cell in ws[header_row] if cell.value]

    print(f"\n  Sheet: {sheet_name}")
    print(f"  Found headers in row {header_row}: {len(headers)} columns")

    # Read all data
    data = []
    total_rows = 0
    analyst_reviewed_rows = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        total_rows += 1

        # Create record dictionary
        record = {}
        for i, header in enumerate(headers):
            if i < len(row):
                record[header] = row[i]

        # Filter for analyst-reviewed records (has Final Decision)
        final_decision = record.get('Final Decision')
        if final_decision and str(final_decision).strip() not in ['', 'nan', 'None']:
            data.append(record)
            analyst_reviewed_rows += 1

    print(f"  Total rows: {total_rows:,}")
    print(f"  Rows with Final Decision: {analyst_reviewed_rows:,} ({analyst_reviewed_rows/total_rows*100:.1f}%)")

    wb.close()

    print(f"\nTotal analyst-reviewed records loaded: {len(data):,}")
    return data

def extract_vendor_patterns(data):
    """Extract vendor-specific patterns from use tax data"""
    vendor_stats = defaultdict(lambda: {
        'transaction_count': 0,
        'refund_bases': Counter(),
        'tax_categories': Counter(),
        'description_keywords': Counter(),
        'final_decisions': Counter()
    })

    for record in data:
        vendor_name = record.get('Vendor Name', '').strip() if record.get('Vendor Name') else ''
        if not vendor_name or vendor_name in ['', 'None', 'nan']:
            continue

        stats = vendor_stats[vendor_name]
        stats['transaction_count'] += 1

        # Refund basis (FILTERED - no PDFs, no invoice numbers)
        refund_basis = record.get('Refund Basis')
        if is_valid_refund_basis(refund_basis):
            stats['refund_bases'][str(refund_basis).strip()] += 1

        # Tax category (FILTERED - no numbers, no headers)
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            stats['tax_categories'][str(tax_cat).strip()] += 1

        # Final Decision (FILTERED - no numbers, no headers, no booleans)
        final_decision = record.get('Final Decision')
        if is_valid_category_value(final_decision):
            stats['final_decisions'][str(final_decision).strip()] += 1

        # Description keywords from Description column
        # FILTERED - no numbers, must have letters, no stopwords
        description = record.get('Description')
        if description and str(description).strip() not in ['', 'nan', 'None']:
            words = str(description).split()
            for word in words:
                if is_valid_keyword(word):
                    stats['description_keywords'][word] += 1

        # Also extract from Add'l Info column
        addl_info = record.get("Add'l Info")
        if addl_info and str(addl_info).strip() not in ['', 'nan', 'None']:
            words = str(addl_info).split()
            for word in words:
                if is_valid_keyword(word):
                    stats['description_keywords'][word] += 1

    # Convert to vendor patterns format (GL accounts removed per user request)
    vendor_patterns = []
    for vendor_name, stats in sorted(vendor_stats.items()):
        # Get most common values
        top_refund_basis = stats['refund_bases'].most_common(1)[0][0] if stats['refund_bases'] else 'Unknown'
        top_tax_category = stats['tax_categories'].most_common(1)[0][0] if stats['tax_categories'] else 'Unknown'
        top_keywords = [word for word, _ in stats['description_keywords'].most_common(10)]
        top_decision = stats['final_decisions'].most_common(1)[0][0] if stats['final_decisions'] else 'Unknown'

        pattern = {
            'vendor_name': vendor_name,
            'tax_type': 'use_tax',  # NEW: Mark as use tax
            'historical_sample_count': stats['transaction_count'],
            'historical_success_rate': calculate_success_rate(stats['final_decisions']),
            'typical_refund_basis': top_refund_basis,
            'typical_final_decision': top_decision,
            'common_tax_categories': [cat for cat, _ in stats['tax_categories'].most_common(3)],
            'common_refund_bases': [rb for rb, _ in stats['refund_bases'].most_common(3)],
            'description_keywords': top_keywords[:5] if top_keywords else []
        }
        vendor_patterns.append(pattern)

    print(f"\nExtracted patterns for {len(vendor_patterns)} vendors")
    return vendor_patterns

def calculate_success_rate(final_decisions):
    """Calculate success rate based on final decisions"""
    if not final_decisions:
        return 1.0

    total = sum(final_decisions.values())
    # Decisions indicating refund
    refund_decisions = sum(count for decision, count in final_decisions.items()
                          if 'add to claim' in str(decision).lower() or 'refund' in str(decision).lower())

    return round(refund_decisions / total, 4) if total > 0 else 1.0

def extract_keyword_patterns(data):
    """Extract keyword patterns across all categories"""
    keywords = {
        'tax_categories': Counter(),
        'refund_basis_terms': Counter(),
        'final_decisions': Counter(),
        'description_keywords': Counter()
    }

    for record in data:
        # Tax categories
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            keywords['tax_categories'][str(tax_cat).strip()] += 1

        # Refund basis terms
        refund_basis = record.get('Refund Basis')
        if is_valid_refund_basis(refund_basis):
            keywords['refund_basis_terms'][str(refund_basis).strip()] += 1

        # Final decisions
        final_decision = record.get('Final Decision')
        if is_valid_category_value(final_decision):
            keywords['final_decisions'][str(final_decision).strip()] += 1

        # Description keywords
        for col in ['Description', "Add'l Info"]:
            desc = record.get(col)
            if desc and str(desc).strip() not in ['', 'nan', 'None']:
                words = str(desc).split()
                for word in words:
                    if is_valid_keyword(word):
                        keywords['description_keywords'][word] += 1

    # Format as list of keyword categories
    keyword_patterns = []
    for category, counter in keywords.items():
        if counter:
            pattern = {
                'category': category,
                'tax_type': 'use_tax',  # NEW: Mark as use tax
                'keywords': [
                    {
                        'keyword': keyword,
                        'frequency': count
                    }
                    for keyword, count in counter.most_common(50)
                ]
            }
            keyword_patterns.append(pattern)

    print(f"\nExtracted {len(keyword_patterns)} keyword categories")
    return keyword_patterns

def extract_refund_basis_patterns(data):
    """Extract refund basis patterns with Final Decision context - FILTERED
    Returns ALL vendors per refund basis (not just 5 examples) per user requirement
    """
    refund_basis_stats = Counter()
    refund_basis_vendors = defaultdict(set)
    refund_basis_decisions = defaultdict(Counter)

    total_records = len(data)

    for record in data:
        refund_basis = record.get('Refund Basis')
        # Apply filter - no PDFs, no invoice numbers, must be text
        if is_valid_refund_basis(refund_basis):
            rb = str(refund_basis).strip()
            refund_basis_stats[rb] += 1

            vendor = record.get('Vendor Name', '').strip() if record.get('Vendor Name') else ''
            if vendor and vendor not in ['None', 'nan', '']:
                refund_basis_vendors[rb].add(vendor)

            final_decision = record.get('Final Decision')
            if final_decision and str(final_decision).strip() not in ['', 'nan', 'None']:
                refund_basis_decisions[rb][str(final_decision).strip()] += 1

    patterns = []
    for refund_basis, count in refund_basis_stats.most_common():
        # Get most common decision for this refund basis
        top_decision = refund_basis_decisions[refund_basis].most_common(1)
        top_decision_text = top_decision[0][0] if top_decision else 'Unknown'

        pattern = {
            'refund_basis': refund_basis,
            'tax_type': 'use_tax',  # NEW: Mark as use tax
            'usage_count': count,
            'percentage': round(count / total_records * 100, 1),
            'typical_final_decision': top_decision_text,
            'all_vendors': sorted(list(refund_basis_vendors[refund_basis])),  # ALL vendors, sorted alphabetically
            'vendor_count': len(refund_basis_vendors[refund_basis])
        }
        patterns.append(pattern)

    print(f"\nExtracted {len(patterns)} refund basis patterns")
    return patterns

def main():
    print("=" * 80)
    print("EXTRACTING PATTERNS FROM PHASE 3 2021 USE TAX FILE")
    print("=" * 80)

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load data
    data = load_use_tax_data(USE_TAX_FILE, sheet_name='2021')

    if not data:
        print("No data loaded. Exiting.")
        return

    # Extract vendor patterns
    print("\n" + "=" * 80)
    print("EXTRACTING VENDOR PATTERNS")
    print("=" * 80)
    vendor_patterns = extract_vendor_patterns(data)

    # Extract keyword patterns
    print("\n" + "=" * 80)
    print("EXTRACTING KEYWORD PATTERNS")
    print("=" * 80)
    keyword_patterns = extract_keyword_patterns(data)

    # Extract refund basis patterns
    print("\n" + "=" * 80)
    print("EXTRACTING REFUND BASIS PATTERNS")
    print("=" * 80)
    refund_basis_patterns = extract_refund_basis_patterns(data)

    # Save all patterns
    print("\n" + "=" * 80)
    print("SAVING PATTERN FILES")
    print("=" * 80)

    # Save vendor patterns
    vendor_output = OUTPUT_DIR / 'vendor_patterns.json'
    with open(vendor_output, 'w', encoding='utf-8') as f:
        json.dump(vendor_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(vendor_patterns)} vendor patterns to {vendor_output}")

    # Save keyword patterns
    keyword_output = OUTPUT_DIR / 'keyword_patterns.json'
    with open(keyword_output, 'w', encoding='utf-8') as f:
        json.dump(keyword_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(keyword_patterns)} keyword categories to {keyword_output}")

    # Save refund basis patterns
    refund_basis_output = OUTPUT_DIR / 'refund_basis_patterns.json'
    with open(refund_basis_output, 'w', encoding='utf-8') as f:
        json.dump(refund_basis_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(refund_basis_patterns)} refund basis patterns to {refund_basis_output}")

    # Save extraction stats
    stats_output = OUTPUT_DIR / 'extraction_stats.json'
    with open(stats_output, 'w', encoding='utf-8') as f:
        json.dump({
            'source_file': 'Phase_3_2021_Use Tax_10-17-25.xlsx',
            'sheet': '2021',
            'total_records_extracted': len(data),
            'vendors_found': len(vendor_patterns),
            'keyword_categories': len(keyword_patterns),
            'refund_bases': len(refund_basis_patterns)
        }, f, indent=2)
    print(f"[OK] Saved extraction statistics to {stats_output}")

    print("\n" + "=" * 80)
    print("COMPLETE!")
    print("=" * 80)
    print(f"\nAnalyst-reviewed records processed: {len(data):,}")
    print(f"Vendors: {len(vendor_patterns)}")
    print(f"Keyword categories: {len(keyword_patterns)}")
    print(f"Refund basis patterns: {len(refund_basis_patterns)}")

if __name__ == '__main__':
    main()
