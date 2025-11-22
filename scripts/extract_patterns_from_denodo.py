"""
Extract patterns from Denodo Excel files (2019-2023) - analyst-reviewed records only
Filters for records where 'Final Decision' column is populated
WITH QUALITY FILTERING - No junk data (PDFs, numeric codes, etc.)
"""
import openpyxl
import json
from collections import defaultdict, Counter
from pathlib import Path
import glob
import sys

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from pattern_filters import (
    is_valid_refund_basis,
    is_valid_keyword,
    is_valid_category_value,
    is_text_gl_description,
    is_meaningful_material_code,
    is_meaningful_location,
    clean_wac_citations
)

# File paths
DENODO_DIR = r'C:\Users\jacob\Desktop\Denodo&UseTax'
OUTPUT_DIR = Path(r'c:\Users\jacob\refundengine\refundengine\extracted_patterns')

# Denodo files to process (skip 2018 due to format error)
DENODO_FILES = [
    '2019 Records in Denodo*.xlsx',
    '2020 Records in Denodo*.xlsx',
    '2021 Records in Denodo*.xlsx',
    '2022 Records in Denodo*.xlsx',
    '2023 Records in Denodo*.xlsx'
]

def find_denodo_files(directory, patterns):
    """Find all Denodo files matching the patterns"""
    files_found = []
    for pattern in patterns:
        matches = glob.glob(str(Path(directory) / pattern))
        files_found.extend(matches)
    return files_found

def load_denodo_data(file_paths):
    """Load and parse Denodo Excel files, filtering for analyst-reviewed records"""
    print(f"Loading Denodo files...")
    all_data = []
    file_stats = {}

    for file_path in file_paths:
        file_name = Path(file_path).name
        print(f"\n{'='*80}")
        print(f"Processing: {file_name}")
        print(f"{'='*80}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                # Skip stats sheets
                if sheet_name.lower() in ['stats', 'summary']:
                    continue

                ws = wb[sheet_name]
                print(f"  Sheet: {sheet_name}")

                # Find header row
                header_row = None
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                    # Look for key Denodo columns
                    if any(cell and ('name1_po_vendor_name' in str(cell) or 'Vendor Name' in str(cell)) for cell in row):
                        header_row = i
                        headers = [cell if cell else f'Column_{idx}' for idx, cell in enumerate(row)]
                        print(f"  Found headers in row {header_row}: {len(headers)} columns")
                        break

                if not header_row:
                    print(f"  WARNING: No header row found in {sheet_name}, skipping...")
                    continue

                # Extract data
                total_rows = 0
                filtered_rows = 0
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not any(row):  # Skip empty rows
                        continue

                    total_rows += 1
                    row_dict = dict(zip(headers, row))

                    # Filter: Must have Final Decision populated
                    final_decision = row_dict.get('Final Decision')
                    if final_decision and str(final_decision).strip() not in ['', 'nan', 'None']:
                        all_data.append({
                            'source_file': file_name,
                            **row_dict
                        })
                        filtered_rows += 1

                print(f"  Total rows: {total_rows:,}")
                print(f"  Rows with Final Decision: {filtered_rows:,} ({filtered_rows/total_rows*100:.1f}%)")

                file_stats[file_name] = {
                    'total_rows': total_rows,
                    'filtered_rows': filtered_rows,
                    'filter_rate': filtered_rows/total_rows*100 if total_rows > 0 else 0
                }

            wb.close()

        except Exception as e:
            print(f"  ERROR processing {file_name}: {str(e)}")
            file_stats[file_name] = {'error': str(e)}

    print(f"\n{'='*80}")
    print(f"LOADING COMPLETE")
    print(f"{'='*80}")
    print(f"Total files processed: {len(file_stats)}")
    print(f"Total analyst-reviewed records loaded: {len(all_data):,}")

    return all_data, file_stats

def extract_vendor_patterns(data):
    """Extract vendor-specific patterns from Denodo data"""
    vendor_stats = defaultdict(lambda: {
        'transaction_count': 0,
        'refund_bases': Counter(),
        'tax_categories': Counter(),
        'product_types': Counter(),
        'final_decisions': Counter(),
        'description_keywords': Counter(),
        'gl_accounts': Counter()
    })

    for record in data:
        # Try multiple vendor name columns
        vendor_name = (record.get('name1_po_vendor_name') or
                      record.get('Vendor Name') or
                      record.get('name1_vendor_name') or '')

        # Convert to string if not already
        vendor_name = str(vendor_name).strip() if vendor_name not in ['', None, 'None', 'nan'] else ''

        if not vendor_name or vendor_name in ['None', 'nan', '']:
            continue

        stats = vendor_stats[vendor_name]
        stats['transaction_count'] += 1

        # Refund basis (FILTERED - no PDFs, no invoice numbers)
        refund_basis = record.get('Refund Basis')
        if is_valid_refund_basis(refund_basis):
            stats['refund_bases'][str(refund_basis).strip()] += 1

        # Tax category (FILTERED - no numbers, no headers)
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            stats['tax_categories'][str(tax_cat).strip()] += 1

        # Product type from material group description (FILTERED)
        product_type = record.get('matk1_po_material_group_desc')
        if is_valid_category_value(product_type):
            stats['product_types'][str(product_type).strip()] += 1

        # Final Decision (FILTERED - no numbers, no headers, no booleans)
        final_decision = record.get('Final Decision')
        if is_valid_category_value(final_decision):
            stats['final_decisions'][str(final_decision).strip()] += 1

        # GL Account - TEXT descriptions only (no numeric codes)
        gl_account = record.get('hkont_account_number')
        if is_text_gl_description(gl_account):
            stats['gl_accounts'][str(gl_account).strip()] += 1

        # Description keywords from Column CC (sgtxt_line_item_text)
        # FILTERED - no numbers, must have letters, no stopwords
        description = record.get('txz01_po_description') or record.get('sgtxt_line_item_text')
        if description and str(description).strip() not in ['', 'nan', 'None', 'See SOW', 'See Excel']:
            words = str(description).split()
            for word in words:
                if is_valid_keyword(word):
                    stats['description_keywords'][word] += 1

        # Also extract from Column BK (post1_wbs_description)
        wbs_desc = record.get('post1_wbs_description')
        if wbs_desc and str(wbs_desc).strip() not in ['', 'nan', 'None', 'See SOW', 'See Excel']:
            words = str(wbs_desc).split()
            for word in words:
                if is_valid_keyword(word):
                    stats['description_keywords'][word] += 1

    # Convert to vendor patterns format (GL accounts removed per user request)
    vendor_patterns = []
    for vendor_name, stats in sorted(vendor_stats.items()):
        # Get most common values
        top_refund_basis = stats['refund_bases'].most_common(1)[0][0] if stats['refund_bases'] else 'Unknown'
        top_tax_category = stats['tax_categories'].most_common(1)[0][0] if stats['tax_categories'] else 'Unknown'
        top_product_type = stats['product_types'].most_common(1)[0][0] if stats['product_types'] else 'Unknown'
        top_keywords = [word for word, _ in stats['description_keywords'].most_common(10)]
        top_decision = stats['final_decisions'].most_common(1)[0][0] if stats['final_decisions'] else 'Unknown'

        pattern = {
            'vendor_name': vendor_name,
            'product_type': top_product_type,
            'historical_sample_count': stats['transaction_count'],
            'historical_success_rate': calculate_success_rate(stats['final_decisions']),
            'typical_refund_basis': top_refund_basis,
            'typical_final_decision': top_decision,
            'common_tax_categories': [cat for cat, _ in stats['tax_categories'].most_common(3)],
            'common_refund_bases': [rb for rb, _ in stats['refund_bases'].most_common(3)],
            'description_keywords': top_keywords[:5] if top_keywords else []
            # GL accounts removed per user request - "I don't care about gl accounts for now"
        }
        vendor_patterns.append(pattern)

    print(f"\nExtracted patterns for {len(vendor_patterns)} vendors")
    return vendor_patterns

def calculate_success_rate(final_decisions):
    """Calculate success rate based on final decisions"""
    if not final_decisions:
        return 1.0

    # Count decisions that indicate approval
    approval_keywords = ['refund', 'approved', 'yes', 'taxable']
    total = sum(final_decisions.values())
    approved = sum(count for decision, count in final_decisions.items()
                   if any(keyword in str(decision).lower() for keyword in approval_keywords))

    return approved / total if total > 0 else 1.0

def extract_keyword_patterns(data):
    """Extract keyword patterns from Denodo data"""
    keywords = {
        'tax_categories': Counter(),
        'product_types': Counter(),
        'refund_basis_terms': Counter(),
        'final_decisions': Counter(),
        'gl_account_descriptions': Counter(),
        'material_groups': Counter()
    }

    for record in data:
        # Tax categories (FILTERED - no numbers, no headers)
        tax_cat = record.get('Tax Category')
        if is_valid_category_value(tax_cat):
            keywords['tax_categories'][str(tax_cat).strip()] += 1

        # Product types from material group (FILTERED)
        product_type = record.get('matk1_po_material_group_desc')
        if is_valid_category_value(product_type):
            keywords['product_types'][str(product_type).strip()] += 1

        # Refund basis (FILTERED - no PDFs, no invoice numbers)
        refund_basis = record.get('Refund Basis')
        if is_valid_refund_basis(refund_basis):
            keywords['refund_basis_terms'][str(refund_basis).strip()] += 1

        # Final decisions (FILTERED - no numbers, no headers, no booleans)
        final_decision = record.get('Final Decision')
        if is_valid_category_value(final_decision):
            keywords['final_decisions'][str(final_decision).strip()] += 1

        # GL Account descriptions (TEXT only, no numeric codes)
        gl_desc = record.get('txt50_account_description')
        if is_text_gl_description(gl_desc):
            keywords['gl_account_descriptions'][str(gl_desc).strip()] += 1

        # Material groups (FILTERED - only codes with letters)
        material_group = record.get('matk1_po_material_group')
        if is_meaningful_material_code(material_group):
            keywords['material_groups'][str(material_group).strip()] += 1

    # Format as structured JSON
    keyword_patterns = []

    for category, counter in keywords.items():
        if counter:
            category_data = {
                'category': category,
                'description': get_category_description(category),
                'keywords': [
                    {'term': term, 'count': count}
                    for term, count in counter.most_common(50)  # Top 50 per category
                ]
            }
            keyword_patterns.append(category_data)

    print(f"\nExtracted {len(keyword_patterns)} keyword categories")
    return keyword_patterns

def get_category_description(category):
    """Get description for keyword category"""
    descriptions = {
        'tax_categories': 'Tax classification categories from analyst reviews',
        'product_types': 'Product/service types from material group descriptions',
        'refund_basis_terms': 'Refund basis categories from analyst decisions',
        'final_decisions': 'Analyst final decision outcomes',
        'gl_account_descriptions': 'General ledger account descriptions',
        'material_groups': 'SAP material group codes'
    }
    return descriptions.get(category, '')

def extract_refund_basis_patterns(data):
    """Extract refund basis patterns with Final Decision context - FILTERED
    Returns ALL vendors per refund basis (not just 5 examples) per user requirement
    """
    refund_basis_stats = Counter()
    refund_basis_vendors = defaultdict(set)
    refund_basis_decisions = defaultdict(Counter)

    total_records = len(data)

    for record in data:
        refund_basis = record.get('Refund Basis')
        # Apply filter - no PDFs, no invoice numbers, must be text
        if is_valid_refund_basis(refund_basis):
            rb = str(refund_basis).strip()
            refund_basis_stats[rb] += 1

            vendor = (record.get('name1_po_vendor_name') or
                     record.get('Vendor Name') or '')
            vendor = str(vendor).strip() if vendor not in ['', None, 'None', 'nan'] else ''
            if vendor and vendor not in ['None', 'nan', '']:
                refund_basis_vendors[rb].add(vendor)

            final_decision = record.get('Final Decision')
            if final_decision and str(final_decision).strip() not in ['', 'nan', 'None']:
                refund_basis_decisions[rb][str(final_decision).strip()] += 1

    patterns = []
    for refund_basis, count in refund_basis_stats.most_common():
        # Get most common decision for this refund basis
        top_decision = refund_basis_decisions[refund_basis].most_common(1)
        top_decision_text = top_decision[0][0] if top_decision else 'Unknown'

        pattern = {
            'refund_basis': refund_basis,
            'usage_count': count,
            'percentage': round(count / total_records * 100, 1),
            'typical_final_decision': top_decision_text,
            'all_vendors': sorted(list(refund_basis_vendors[refund_basis])),  # ALL vendors, sorted alphabetically
            'vendor_count': len(refund_basis_vendors[refund_basis])
        }
        patterns.append(pattern)

    print(f"\nExtracted {len(patterns)} refund basis patterns")
    return patterns

def extract_citation_patterns(data):
    """
    Extract legal citation patterns - ONLY IF REAL EXAMPLES EXIST
    Focus on TEXT descriptions, NO numeric codes per user requirement
    """
    import re

    wac_examples = []
    wac_pattern = re.compile(r'WAC\s+\d{3}-\d{2}-\d{3,4}[A-Z]?', re.IGNORECASE)
    gl_descriptions = Counter()  # TEXT descriptions only, not numeric codes

    for record in data:
        # Look for WAC patterns in all text fields - must match actual pattern
        for key, value in record.items():
            if value and isinstance(value, str):
                if wac_pattern.search(value):
                    wac_examples.append(value)

        # GL Account - TEXT descriptions only (user wants NO numeric codes)
        gl_desc = record.get('txt50_account_description')
        if is_text_gl_description(gl_desc):
            gl_descriptions[str(gl_desc).strip()] += 1

    # Clean WAC citations - only keep actual matches
    wac_examples = clean_wac_citations(wac_examples)

    patterns = []

    # Only add WAC if we found real examples
    if wac_examples:
        patterns.append({
            'citation_type': 'WAC',
            'pattern': r'WAC \d{3}-\d{2}-\d{3,4}[A-Z]?',
            'examples': wac_examples[:10],
            'description': 'Washington Administrative Code tax regulations'
        })

    # Only add GL descriptions if we have text-based ones (no numeric codes)
    if gl_descriptions:
        patterns.append({
            'citation_type': 'GL_Account_Descriptions',
            'description': 'General Ledger account TEXT descriptions (numeric codes excluded per user requirement)',
            'descriptions': [
                {'description': desc, 'frequency': freq}
                for desc, freq in gl_descriptions.most_common(20)
            ]
        })

    # RCW removed - no examples found, user said don't force patterns

    print(f"\nExtracted citation patterns: {len(wac_examples)} WAC examples, {len(gl_descriptions)} GL descriptions (TEXT only)")
    return patterns

def merge_with_existing_vendors(new_patterns, existing_file):
    """Merge new vendor patterns with existing ones"""
    if not existing_file.exists():
        print("No existing vendor patterns file found, using new data only")
        return new_patterns

    with open(existing_file, 'r', encoding='utf-8') as f:
        existing_patterns = json.load(f)

    # Create lookup by vendor name
    existing_by_name = {p['vendor_name']: p for p in existing_patterns}

    merged = []
    new_vendors = 0
    updated_vendors = 0

    for new_pattern in new_patterns:
        vendor_name = new_pattern['vendor_name']

        if vendor_name in existing_by_name:
            # Merge: add transaction counts
            existing = existing_by_name[vendor_name]
            existing['historical_sample_count'] += new_pattern['historical_sample_count']

            # Update with newer data if available
            if new_pattern.get('typical_refund_basis') != 'Unknown':
                existing['typical_refund_basis'] = new_pattern['typical_refund_basis']

            # Add/update new fields from Denodo
            if new_pattern.get('typical_final_decision'):
                existing['typical_final_decision'] = new_pattern['typical_final_decision']

            if new_pattern.get('common_tax_categories'):
                existing['common_tax_categories'] = new_pattern['common_tax_categories']

            if new_pattern.get('common_refund_bases'):
                existing['common_refund_bases'] = new_pattern['common_refund_bases']

            # GL accounts removed per user request

            merged.append(existing)
            updated_vendors += 1
            del existing_by_name[vendor_name]
        else:
            # New vendor from Denodo
            merged.append(new_pattern)
            new_vendors += 1

    # Add remaining existing vendors that weren't in Denodo data
    merged.extend(existing_by_name.values())

    # Sort by transaction count descending
    merged.sort(key=lambda x: x['historical_sample_count'], reverse=True)

    print(f"\n{'='*80}")
    print(f"VENDOR PATTERN MERGE SUMMARY")
    print(f"{'='*80}")
    print(f"New vendors from Denodo: {new_vendors}")
    print(f"Updated existing vendors: {updated_vendors}")
    print(f"Unchanged existing vendors: {len(existing_by_name)}")
    print(f"Total vendors after merge: {len(merged)}")

    return merged

def merge_keyword_patterns(new_patterns, existing_file):
    """Merge new keyword patterns with existing ones"""
    if not existing_file.exists():
        print("No existing keyword patterns file found, using new data only")
        return new_patterns

    with open(existing_file, 'r', encoding='utf-8') as f:
        existing_patterns = json.load(f)

    # Create lookup by category
    existing_by_category = {p['category']: p for p in existing_patterns}

    merged = []

    for new_pattern in new_patterns:
        category = new_pattern['category']

        if category in existing_by_category:
            # Merge keywords
            existing = existing_by_category[category]
            existing_keywords = {kw['term']: kw['count'] for kw in existing['keywords']}

            for new_kw in new_pattern['keywords']:
                term = new_kw['term']
                if term in existing_keywords:
                    existing_keywords[term] += new_kw['count']
                else:
                    existing_keywords[term] = new_kw['count']

            # Rebuild keywords list sorted by count
            existing['keywords'] = [
                {'term': term, 'count': count}
                for term, count in sorted(existing_keywords.items(), key=lambda x: x[1], reverse=True)
            ][:50]  # Keep top 50

            merged.append(existing)
            del existing_by_category[category]
        else:
            # New category
            merged.append(new_pattern)

    # Add remaining existing categories
    merged.extend(existing_by_category.values())

    return merged

def main():
    """Main execution"""
    print("=" * 80)
    print("EXTRACTING PATTERNS FROM DENODO FILES (ANALYST-REVIEWED RECORDS)")
    print("=" * 80)

    # Find Denodo files
    denodo_files = find_denodo_files(DENODO_DIR, DENODO_FILES)
    print(f"\nFound {len(denodo_files)} Denodo files:")
    for f in denodo_files:
        print(f"  - {Path(f).name}")

    if not denodo_files:
        print("ERROR: No Denodo files found")
        return

    # Load data (filtered for Final Decision)
    data, file_stats = load_denodo_data(denodo_files)

    if not data:
        print("ERROR: No analyst-reviewed records found")
        return

    # Extract patterns
    print("\n" + "=" * 80)
    print("EXTRACTING VENDOR PATTERNS")
    print("=" * 80)
    vendor_patterns = extract_vendor_patterns(data)

    # Merge with existing
    vendor_file = OUTPUT_DIR / 'vendor_patterns.json'
    merged_vendors = merge_with_existing_vendors(vendor_patterns, vendor_file)

    print("\n" + "=" * 80)
    print("EXTRACTING KEYWORD PATTERNS")
    print("=" * 80)
    keyword_patterns = extract_keyword_patterns(data)

    # Merge with existing
    keyword_file = OUTPUT_DIR / 'keyword_patterns.json'
    merged_keywords = merge_keyword_patterns(keyword_patterns, keyword_file)

    print("\n" + "=" * 80)
    print("EXTRACTING REFUND BASIS PATTERNS")
    print("=" * 80)
    refund_basis_patterns = extract_refund_basis_patterns(data)

    # Citation patterns removed per user request - "I don't need the citations pattern"

    # Save all patterns
    print("\n" + "=" * 80)
    print("SAVING PATTERN FILES")
    print("=" * 80)

    # Save vendor patterns
    vendor_output = OUTPUT_DIR / 'vendor_patterns.json'
    with open(vendor_output, 'w', encoding='utf-8') as f:
        json.dump(merged_vendors, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(merged_vendors)} vendor patterns to {vendor_output}")

    # Save keyword patterns
    keyword_output = OUTPUT_DIR / 'keyword_patterns.json'
    with open(keyword_output, 'w', encoding='utf-8') as f:
        json.dump(merged_keywords, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(merged_keywords)} keyword categories to {keyword_output}")

    # Merge and save refund basis patterns
    existing_rb_file = OUTPUT_DIR / 'refund_basis_patterns.json'
    if existing_rb_file.exists():
        with open(existing_rb_file, 'r', encoding='utf-8') as f:
            existing_rb = json.load(f)

        # Merge by refund basis
        rb_by_name = {rb['refund_basis']: rb for rb in existing_rb}
        for new_rb in refund_basis_patterns:
            rb_name = new_rb['refund_basis']
            if rb_name in rb_by_name:
                rb_by_name[rb_name]['usage_count'] += new_rb['usage_count']
                # Update decision if available
                if new_rb.get('typical_final_decision'):
                    rb_by_name[rb_name]['typical_final_decision'] = new_rb['typical_final_decision']
            else:
                rb_by_name[rb_name] = new_rb

        refund_basis_patterns = list(rb_by_name.values())
        refund_basis_patterns.sort(key=lambda x: x['usage_count'], reverse=True)

    refund_basis_output = OUTPUT_DIR / 'refund_basis_patterns.json'
    with open(refund_basis_output, 'w', encoding='utf-8') as f:
        json.dump(refund_basis_patterns, f, indent=2, ensure_ascii=False)
    print(f"[OK] Saved {len(refund_basis_patterns)} refund basis patterns to {refund_basis_output}")

    # Save extraction stats
    stats_output = OUTPUT_DIR / 'denodo_extraction_stats.json'
    with open(stats_output, 'w', encoding='utf-8') as f:
        json.dump({
            'files_processed': file_stats,
            'total_records_extracted': len(data),
            'vendors_found': len(vendor_patterns),
            'keyword_categories': len(keyword_patterns),
            'refund_bases': len(refund_basis_patterns)
        }, f, indent=2)
    print(f"[OK] Saved extraction statistics to {stats_output}")

    print("\n" + "=" * 80)
    print("COMPLETE!")
    print("=" * 80)
    print(f"\nAnalyst-reviewed records processed: {len(data):,}")
    print(f"Vendors: {len(merged_vendors)}")
    print(f"Keyword categories: {len(merged_keywords)}")
    print(f"Refund basis patterns: {len(refund_basis_patterns)}")

if __name__ == '__main__':
    main()
