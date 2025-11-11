#!/usr/bin/env python3
"""
Export Knowledge Base Metadata to Excel

Exports all document and chunk metadata from Supabase to a single Excel file with multiple sheets.

Usage:
    python scripts/export_metadata_excel.py
    python scripts/export_metadata_excel.py --output metadata.xlsx
"""

import os
from datetime import datetime
from pathlib import Path
from supabase import create_client
from dotenv import load_dotenv
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load environment
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")


def style_header(ws):
    """Apply styling to header row"""
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(output_path: Path):
    """Export all metadata to a single Excel file with multiple sheets"""

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    print("=" * 80)
    print("📤 EXPORTING KNOWLEDGE BASE TO EXCEL")
    print("=" * 80)

    # ========================================================================
    # Sheet 1: Documents
    # ========================================================================
    print("\n📥 Fetching documents...")
    docs_result = (
        supabase.table("knowledge_documents").select("*").order("created_at").execute()
    )

    if docs_result.data:
        ws_docs = wb.create_sheet("Documents")

        # Headers (includes new metadata fields)
        headers = [
            "id",
            "document_type",
            "title",
            "citation",
            "law_category",
            "effective_date",
            "topic_tags",
            "tax_types",
            "industries",
            "referenced_statutes",
            "vendor_name",
            "vendor_category",
            "industry",
            "business_model",
            "primary_products",
            "typical_delivery",
            "tax_notes",
            "confidence_score",
            "data_source",
            "source_file",
            "total_chunks",
            "processing_status",
            "created_at",
            "updated_at",
        ]
        ws_docs.append(headers)

        # Data
        for doc in docs_result.data:
            row = []
            for h in headers:
                value = doc.get(h)
                # Convert arrays to comma-separated strings
                if isinstance(value, list):
                    value = ", ".join(value) if value else ""
                row.append(value)
            ws_docs.append(row)

        # Style
        style_header(ws_docs)
        auto_adjust_columns(ws_docs)

        # Freeze header row
        ws_docs.freeze_panes = "A2"

        print(f"✅ Exported {len(docs_result.data)} documents")
    else:
        print("⚠️  No documents found")

    # ========================================================================
    # Sheet 2: Tax Law Chunks
    # ========================================================================
    print("\n📥 Fetching tax law chunks...")
    try:
        chunks_result = (
            supabase.table("tax_law_chunks")
            .select(
                "id, document_id, chunk_number, citation, section_title, law_category, topic_tags, tax_types, industries, referenced_statutes"
            )
            .order("document_id, chunk_number")
            .execute()
        )

        if chunks_result.data:
            ws_chunks = wb.create_sheet("Tax Law Chunks")

            # Headers (includes new metadata fields)
            headers = [
                "id",
                "document_id",
                "chunk_number",
                "citation",
                "section_title",
                "law_category",
                "topic_tags",
                "tax_types",
                "industries",
                "referenced_statutes",
            ]
            ws_chunks.append(headers)

            # Data
            for chunk in chunks_result.data:
                row = []
                for h in headers:
                    value = chunk.get(h)
                    # Convert arrays to comma-separated strings
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    row.append(value)
                ws_chunks.append(row)

            # Style
            style_header(ws_chunks)
            auto_adjust_columns(ws_chunks)
            ws_chunks.freeze_panes = "A2"

            print(f"✅ Exported {len(chunks_result.data)} tax law chunks")
        else:
            print("⚠️  No tax law chunks found")
    except Exception as e:
        print(f"⚠️  Could not fetch tax law chunks: {e}")

    # ========================================================================
    # Sheet 3: Vendor Background Chunks
    # ========================================================================
    print("\n📥 Fetching vendor background chunks...")
    try:
        vendor_result = (
            supabase.table("vendor_background_chunks")
            .select(
                "id, document_id, chunk_number, vendor_name, vendor_category, document_category"
            )
            .order("document_id, chunk_number")
            .execute()
        )

        if vendor_result.data:
            ws_vendor = wb.create_sheet("Vendor Chunks")

            # Headers
            headers = [
                "id",
                "document_id",
                "chunk_number",
                "vendor_name",
                "vendor_category",
                "document_category",
            ]
            ws_vendor.append(headers)

            # Data
            for chunk in vendor_result.data:
                row = []
                for h in headers:
                    value = chunk.get(h)
                    # Convert arrays to comma-separated strings
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    row.append(value)
                ws_vendor.append(row)

            # Style
            style_header(ws_vendor)
            auto_adjust_columns(ws_vendor)
            ws_vendor.freeze_panes = "A2"

            print(f"✅ Exported {len(vendor_result.data)} vendor chunks")
        else:
            print("⚠️  No vendor chunks found (table is empty)")
    except Exception as e:
        print(f"⚠️  Could not fetch vendor chunks: {e}")

    # ========================================================================
    # Sheet 4: Instructions
    # ========================================================================
    ws_help = wb.create_sheet("📖 Instructions", 0)  # Make it first sheet

    instructions = [
        ["Knowledge Base Metadata Editor"],
        [""],
        ["HOW TO EDIT"],
        ["1. Go to the 'Documents' or 'Tax Law Chunks' sheet"],
        [
            "2. Edit any fields EXCEPT 'id' and 'document_id' (these are used for matching)"
        ],
        ["3. Save this Excel file"],
        ["4. Run: python scripts/import_metadata_excel.py --file <this-file.xlsx>"],
        [""],
        ["EDITABLE FIELDS"],
        [""],
        ["Documents Sheet:"],
        ["  • title - Document title"],
        ["  • citation - Legal citation (e.g., WAC 458-20-15502)"],
        [
            "  • law_category - Category tag (software, digital_goods, manufacturing, etc.)"
        ],
        ["  • effective_date - When the law became effective (YYYY-MM-DD)"],
        [
            "  • topic_tags - Comma-separated topics (e.g., 'digital products, exemptions')"
        ],
        ["  • tax_types - Comma-separated types (e.g., 'sales tax, use tax')"],
        ["  • industries - Comma-separated industries (e.g., 'retail, technology')"],
        ["  • referenced_statutes - Comma-separated laws (e.g., 'RCW 82.04.215')"],
        ["  • vendor_name - For vendor documents only"],
        ["  • vendor_category - Type of vendor"],
        [""],
        ["Tax Law Chunks Sheet:"],
        ["  • citation - Chunk-specific citation"],
        ["  • section_title - Section name or page number"],
        ["  • law_category - Category tag for filtering"],
        [
            "  • topic_tags, tax_types, industries, referenced_statutes - Same as Documents sheet"
        ],
        ["  • Note: Editing Documents sheet will cascade changes to all chunks!"],
        [""],
        ["SUGGESTED CATEGORIES"],
        ["  • software - Computer software taxation"],
        ["  • digital_goods - Digital products"],
        ["  • manufacturing - Manufacturing exemptions"],
        ["  • saas - Software as a Service"],
        ["  • cloud - Cloud computing"],
        ["  • exemption - General exemptions"],
        ["  • procedure - Tax procedures"],
        [""],
        ["IMPORTANT NOTES"],
        ["  • Don't delete the 'id' or 'document_id' columns"],
        ["  • Don't delete rows (won't delete from database)"],
        ["  • The import script will show you a diff before applying changes"],
        ["  • Embeddings and chunk text are preserved"],
        [""],
        ["AFTER EDITING"],
        ["Run: python scripts/import_metadata_excel.py --file metadata.xlsx"],
        [""],
        ["The import will:"],
        ["  1. Show you exactly what changed"],
        ["  2. Ask for confirmation"],
        ["  3. Update only the changed fields in Supabase"],
    ]

    for row in instructions:
        ws_help.append(row)

    # Style instructions
    ws_help["A1"].font = Font(bold=True, size=14, color="4472C4")
    ws_help["A3"].font = Font(bold=True, size=12)
    ws_help["A9"].font = Font(bold=True, size=12)
    ws_help["A26"].font = Font(bold=True, size=12)
    ws_help["A37"].font = Font(bold=True, size=12)

    ws_help.column_dimensions["A"].width = 80

    # ========================================================================
    # Save workbook
    # ========================================================================
    wb.save(output_path)

    print("\n" + "=" * 80)
    print("✅ EXPORT COMPLETE")
    print("=" * 80)
    print(f"\n📁 Excel file created: {output_path.absolute()}")
    print("\n📝 Next steps:")
    print("  1. Open the Excel file")
    print("  2. Edit metadata in the 'Documents' or 'Tax Law Chunks' sheets")
    print("  3. Save the file")
    print("  4. Run: python scripts/import_metadata_excel.py --file <your-file.xlsx>")
    print()


def main():
    parser = argparse.ArgumentParser(
        description="Export knowledge base metadata to Excel"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file path (default: metadata_exports/metadata_TIMESTAMP.xlsx)",
    )

    args = parser.parse_args()

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_dir = Path("./metadata_exports")
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"metadata_{timestamp}.xlsx"

    try:
        export_to_excel(output_path)
        return 0
    except Exception as e:
        print(f"\n❌ Error during export: {e}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
