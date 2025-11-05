#!/usr/bin/env python3
"""
DOR Filing Worksheet Generator

Creates DOR-compliant filing worksheet.

Usage:
    python scripts/generate_dor_filing.py --client_id 1
"""

import argparse
import sys
import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
import excel_utils as xl

def get_database_connection():
    project_root = Path(__file__).parent.parent
    db_path = project_root / "database" / "refund_engine.db"
    return sqlite3.connect(str(db_path))

def generate_dor_filing_worksheet(client_id):
    """Generate DOR filing worksheet."""
    conn = get_database_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT client_name FROM clients WHERE id = ?", (client_id,))
    client = cursor.fetchone()
    if not client:
        print(f"❌ Client {client_id} not found")
        return None

    client_name = client[0]
    print(f"\n📋 Generating DOR filing worksheet for: {client_name}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Claim Summary"

    xl.add_title(ws, "Washington State DOR - Refund Claim", row=1, end_col=8)
    ws.cell(2, 1, f"Taxpayer: {client_name}")
    ws.cell(3, 1, f"Date: {datetime.now().strftime('%B %d, %Y')}")

    headers = ["Invoice #", "Date", "Vendor", "Description", "Amount", "Tax Paid", "Refund", "Statute"]
    xl.add_header_row(ws, headers, row=5)

    cursor.execute("""
        SELECT i.invoice_number, i.invoice_date, i.vendor_name, li.item_description,
               li.line_total, li.sales_tax_on_line, ra.estimated_refund_amount,
               ra.criteria_matching_json
        FROM refund_analysis ra
        JOIN invoices i ON ra.invoice_id = i.id
        JOIN invoice_line_items li ON ra.line_item_id = li.id
        WHERE i.client_id = ? AND ra.potentially_eligible = 1
    """, (client_id,))

    row = 6
    for rec in cursor.fetchall():
        for col, val in enumerate(rec, 1):
            ws.cell(row, col, val)
        row += 1

    xl.format_as_table(ws, start_row=5, end_row=row-1, end_col=8)

    project_root = Path(__file__).parent.parent
    output_dir = project_root / "outputs" / "dor_filings"
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{client_name.replace(' ', '_')}_DOR_Filing_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename

    wb.save(filepath)
    conn.close()

    print(f"✅ DOR filing worksheet generated: {filepath}")
    return str(filepath)

def main():
    parser = argparse.ArgumentParser(description="Generate DOR filing worksheet")
    parser.add_argument('--client_id', required=True, type=int)
    args = parser.parse_args()

    generate_dor_filing_worksheet(args.client_id)
    return 0

if __name__ == "__main__":
    sys.exit(main())
