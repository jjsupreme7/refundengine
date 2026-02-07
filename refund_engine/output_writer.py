from __future__ import annotations

from pathlib import Path
import shutil
from typing import Any

from openpyxl import Workbook, load_workbook

from refund_engine.constants import AI_OUTPUT_COLUMNS
from refund_engine.datasets import DatasetConfig, read_source_dataframe


def ensure_output_file(config: DatasetConfig):
    if config.output_file.exists():
        return

    config.output_file.parent.mkdir(parents=True, exist_ok=True)

    if config.source_file.exists() and config.source_file.suffix.lower() == ".xlsx":
        shutil.copy2(config.source_file, config.output_file)
        return

    # Fallback: create a basic xlsx from source dataframe
    source_df = read_source_dataframe(config)
    wb = Workbook()
    ws = wb.active
    if config.sheet_name:
        ws.title = config.sheet_name

    headers = list(source_df.columns)
    ws.append(headers)
    for _, row in source_df.iterrows():
        ws.append(list(row.values))

    wb.save(config.output_file)


def _header_map(worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=col_idx).value
        if value:
            mapping[str(value)] = col_idx
    return mapping


def _ensure_headers(worksheet, headers: tuple[str, ...]) -> dict[str, int]:
    mapping = _header_map(worksheet)
    next_col = worksheet.max_column + 1
    for header in headers:
        if header in mapping:
            continue
        worksheet.cell(row=1, column=next_col, value=header)
        mapping[header] = next_col
        next_col += 1
    return mapping


def _safe_cell_value(value: Any):
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def apply_updates_to_output(
    config: DatasetConfig,
    updates_by_row_index: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    ensure_output_file(config)

    if config.output_file.suffix.lower() != ".xlsx":
        raise ValueError(
            f"Output writing currently supports .xlsx only (got {config.output_file})"
        )

    wb = load_workbook(config.output_file)
    ws = wb[config.sheet_name] if config.sheet_name and config.sheet_name in wb.sheetnames else wb.active

    headers = _ensure_headers(ws, AI_OUTPUT_COLUMNS)
    updated_rows = 0
    skipped_rows: list[int] = []

    for source_index, row_values in updates_by_row_index.items():
        row_number = int(source_index) + 2
        if row_number > ws.max_row:
            skipped_rows.append(int(source_index))
            continue

        for column_name, value in row_values.items():
            if column_name not in headers:
                continue
            ws.cell(
                row=row_number,
                column=headers[column_name],
                value=_safe_cell_value(value),
            )
        updated_rows += 1

    wb.save(config.output_file)
    return {
        "updated_rows": updated_rows,
        "skipped_rows": skipped_rows,
        "output_file": str(config.output_file),
    }
